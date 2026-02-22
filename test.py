# app.py
# -*- coding: utf-8 -*-
"""
叠纸心意旗舰店｜退运费智能审核系统 Web App
技术栈：Streamlit + Pandas + DashScope（通义千问-VL）

Tab1：班牛《退运费自助登记表》清洗 + 规则初筛（金额/账号/实名/物流单号）
Tab2：步骤二入库匹配（正常表 + 已入库物流单号表） + 步骤三AI 视觉复核（多图识别）

【关键防坑点】
1) 读取后第一时间 df.columns strip（必须）
2) st.dataframe 不指定列名，避免 KeyError
3) 缺少必要列 st.error 友好提示
4) Excel 超链接：pandas 默认读不到 hyperlink.target
   - 读取时用 openpyxl 抽取 URL（兼容：超链接对象 / HYPERLINK() 公式 / tooltip / 批注）
     存入辅助列：{列名}__hyperlink（并按 df 行数对齐）
   - 导出时用 openpyxl 写回 hyperlink（保留原文字，如“预览/浏览”，但可点击）
5) 预览链接（exportFilePreview?url=...__...）包含多张图：AI 阶段会拆出多张图片一起送模型查找运费截图
"""

import os
import re
import json
import time
import math
import zipfile
import threading
import traceback
from collections import Counter
from functools import lru_cache
from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any
from urllib.parse import unquote, urlparse, parse_qs

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ====== DashScope（阿里云百炼/通义千问）======
try:
    import dashscope
    from http import HTTPStatus
except Exception:
    dashscope = None
    HTTPStatus = None


# =============================================================================
# 【1】全局可配置变量（列名变化只改这里）
# =============================================================================

COL_AMOUNT_CANDIDATES = [
    "*寄回快递实付金额", "寄回快递实付金额", "*寄回运费金额", "寄回运费金额", "退回运费金额", "*退回运费金额"
]
COL_ALIPAY_ACCOUNT_CANDIDATES = [
    "*退运费的支付宝账号", "退运费的支付宝账号", "支付宝账号", "收款支付宝账号", "支付宝收款账号"
]
COL_ALIPAY_NAME_CANDIDATES = [
    "*退运费的支付宝实名", "退运费的支付宝实名", "支付宝实名", "收款人姓名", "收款人"
]
COL_LOGISTICS_NO_CANDIDATES = [
    "*寄回换货快递单号", "寄回换货快递单号", "*退回物流单号", "退回物流单号", "寄回物流单号", "快递单号"
]
COL_SCREENSHOT_CANDIDATES = [
    "*商品瑕疵+金额截图", "商品瑕疵+金额截图", "寄回运费截图", "运费截图", "截图", "图片URL", "图片链接"
]
COL_ID_CANDIDATES = [
    "ID", "id", "*ID", "旺旺ID", "*旺旺ID", "用户ID", "买家ID", "会员ID"
]
COL_ORDER_NO_CANDIDATES = [
    "订单号", "*订单号", "订单编号", "主订单号", "子订单号", "多笔订单号",
    "订单号（多笔订单分开提交）", "*订单号（多笔订单分开提交）"
]

MAX_REFUND_AMOUNT = 12.0

REGEX_PHONE = re.compile(r"^1[3-9]\d{9}$")
REGEX_EMAIL = re.compile(r"^[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+$")

# ✅ 调整：实名 2~5 个汉字
REGEX_CN_NAME = re.compile(r"^[\u4e00-\u9fa5]{2,5}$")

# ✅ 调整：物流单号 10~16 位字母数字，且必须包含数字
REGEX_LOGISTICS = re.compile(r"^(?=.*\d)[A-Za-z0-9]{10,16}$")
REGEX_MONEY_CLEAN = re.compile(r"[^0-9.\-]")
REGEX_NON_ALNUM = re.compile(r"[^A-Za-z0-9]")
REGEX_URL_IN_PARENS = re.compile(r"\((https?://[^\s)]+)\)")
REGEX_URL_GENERIC = re.compile(r"(https?://[^\s\]\"')]+)")
REGEX_PREVIEW_SPLIT = re.compile(r"__|;|\s+")
REGEX_SCI_NUMBER = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?\d+$")
REGEX_EXCEL_HYPERLINK_FORMULA = re.compile(r'HYPERLINK\(\s*"([^"]+)"\s*[,;]\s*', re.IGNORECASE)
REGEX_EXCEL_URL_FALLBACK = re.compile(r"(https?://[^\s\"')]+)")
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".bmp")
IDENTIFIER_COLUMN_KEYWORDS = (
    "订单", "单号", "物流", "快递", "账号", "支付宝", "流水", "编号", "ID", "id"
)

COL_ABNORMAL_REASON = "异常原因"
COL_AI_EXTRACTED_AMOUNT = "AI提取运费金额"
COL_AI_MATCH = "AI是否一致"
COL_AI_NOTE = "AI异常说明"

# ✅ 入库匹配新增列
COL_INBOUND_FLAG = "是否已入库"     # 值：已入库 / 空
COL_INBOUND_NOTE = "入库匹配说明"   # 值：匹配到已入库表 / 空

DEFAULT_VL_MODEL = "qwen-vl-plus"
PROGRESS_UPDATE_EVERY = 1

# Excel 超链接辅助列后缀（内部使用，导出不会带出去）
HYPERLINK_SUFFIX = "__hyperlink"
HISTORY_FILE_NAME = "operation_history.jsonl"
ARTIFACT_DIR_NAME = "operation_artifacts"
TASK_DIR_NAME = "operation_tasks"

COL_AI_TASK_STATUS = "AI任务状态"
COL_AI_TASK_UPDATED_AT = "AI处理时间"

AI_TASK_STATUS_RUNNING = "running"
AI_TASK_STATUS_PAUSED = "paused"
AI_TASK_STATUS_COMPLETED = "completed"
AI_TASK_STATUS_ERROR = "error"

AI_ROW_STATUS_PENDING = "待处理"
AI_ROW_STATUS_DONE = "已处理"
AI_ROW_STATUS_OUT_OF_SCOPE = "超出本次处理上限"

_AI_TASK_FILE_LOCK = threading.Lock()


# =============================================================================
# 【2】基础工具函数（读取、清洗、校验）
# =============================================================================

def now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_history_file_path() -> Path:
    return Path.cwd() / HISTORY_FILE_NAME


def get_artifact_root_path() -> Path:
    root = Path.cwd() / ARTIFACT_DIR_NAME
    root.mkdir(parents=True, exist_ok=True)
    return root


def _sanitize_file_name(name: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "_", str(name)).strip()
    return safe[:180] if safe else f"{now_ts()}_unnamed.xlsx"


def _extract_display_name_from_artifact_name(name: str) -> str:
    parts = str(name).split("__", 2)
    return parts[2] if len(parts) == 3 else str(name)


def save_artifact_bytes(stage_key: str, file_name: str, data: bytes) -> str:
    if not data:
        return ""
    now = datetime.now()
    date_dir = get_artifact_root_path() / now.strftime("%Y") / now.strftime("%m") / now.strftime("%d")
    date_dir.mkdir(parents=True, exist_ok=True)

    safe_stage = re.sub(r"[^A-Za-z0-9_-]+", "_", str(stage_key)).strip("_") or "stage"
    safe_file = _sanitize_file_name(file_name)
    ts = now.strftime("%Y%m%d_%H%M%S")

    candidate = date_dir / f"{ts}__{safe_stage}__{safe_file}"
    idx = 1
    stem = Path(safe_file).stem
    suffix = Path(safe_file).suffix
    while candidate.exists():
        candidate = date_dir / f"{ts}__{safe_stage}__{stem}_{idx}{suffix}"
        idx += 1

    with open(candidate, "wb") as f:
        f.write(data)
    return candidate.relative_to(Path.cwd()).as_posix()


def load_artifact_catalog_df() -> pd.DataFrame:
    root = get_artifact_root_path()
    rows: List[Dict[str, Any]] = []
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        stat = p.stat()
        mtime = datetime.fromtimestamp(stat.st_mtime)
        rel = p.relative_to(Path.cwd()).as_posix()
        name = p.name
        parts = name.split("__", 2)
        stage_key = ""
        try:
            dt = datetime.strptime(parts[0], "%Y%m%d_%H%M%S") if len(parts) >= 1 else mtime
        except Exception:
            dt = mtime
        if len(parts) >= 2:
            stage_key = parts[1]
        rows.append({
            "timestamp": dt.strftime("%Y-%m-%d %H:%M:%S"),
            "year": dt.strftime("%Y"),
            "month": dt.strftime("%m"),
            "day": dt.strftime("%d"),
            "stage_key": stage_key,
            "file_name": _extract_display_name_from_artifact_name(name),
            "file_path": rel,
            "size_kb": round(stat.st_size / 1024, 2),
        })
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values("timestamp", ascending=False).reset_index(drop=True)


def get_task_root_path() -> Path:
    root = Path.cwd() / TASK_DIR_NAME
    root.mkdir(parents=True, exist_ok=True)
    return root


def _make_ai_task_id() -> str:
    return f"ai_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"


def _get_ai_task_dir(task_id: str) -> Path:
    return get_task_root_path() / str(task_id)


def _get_ai_task_meta_path(task_id: str) -> Path:
    return _get_ai_task_dir(task_id) / "meta.json"


def _get_ai_task_df_path(task_id: str) -> Path:
    return _get_ai_task_dir(task_id) / "df_work.pkl"


def _get_ai_task_source_df_path(task_id: str) -> Path:
    return _get_ai_task_dir(task_id) / "source_df.pkl"


def init_ai_task_dataframe(df_source: pd.DataFrame, total_rows: int) -> pd.DataFrame:
    df = df_source.reset_index(drop=True).copy()
    total = max(0, min(int(total_rows), len(df)))

    # 每次新建任务都重置 AI 结果列，避免使用旧结果误判。
    df[COL_AI_EXTRACTED_AMOUNT] = None
    df[COL_AI_MATCH] = None
    df[COL_AI_NOTE] = ""
    df[COL_AI_TASK_STATUS] = AI_ROW_STATUS_PENDING
    df[COL_AI_TASK_UPDATED_AT] = ""

    if total < len(df):
        out_scope_idx = df.index[total:]
        df.loc[out_scope_idx, COL_AI_TASK_STATUS] = AI_ROW_STATUS_OUT_OF_SCOPE
        df.loc[out_scope_idx, COL_AI_NOTE] = "未处理（超过本次最大处理行数限制）"

    return df


def create_ai_task_state(
    df_source: pd.DataFrame,
    source_file: str,
    col_amount: str,
    col_shot: str,
    total_rows: int,
    model_name: str,
    max_images: int,
    min_interval_sec: float,
    max_retries: int,
    backoff_base_sec: float,
) -> Dict[str, Any]:
    src = df_source.reset_index(drop=True).copy()
    total = max(0, min(int(total_rows), len(src)))
    task_id = _make_ai_task_id()

    task = {
        "task_id": task_id,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "finished_at": "",
        "status": AI_TASK_STATUS_PAUSED,
        "source_file": source_file,
        "input_rows": len(src),
        "total": total,
        "next_idx": 0,
        "col_amount": col_amount,
        "col_shot": col_shot,
        "model_name": model_name,
        "max_images": int(max_images),
        "min_interval_sec": float(min_interval_sec),
        "max_retries": int(max_retries),
        "backoff_base_sec": float(backoff_base_sec),
        "history_logged": False,
        "alignment_report": None,
        "error_message": "",
        "artifacts": [],
        "df_work": init_ai_task_dataframe(src, total),
        "source_df": src,
    }
    save_ai_task_state(task)
    return task


def save_ai_task_state(task: Dict[str, Any]) -> None:
    task_id = str(task.get("task_id", "")).strip()
    if not task_id:
        return

    task_dir = _get_ai_task_dir(task_id)
    task_dir.mkdir(parents=True, exist_ok=True)

    with _AI_TASK_FILE_LOCK:
        df_work = task.get("df_work")
        if isinstance(df_work, pd.DataFrame):
            df_work.to_pickle(_get_ai_task_df_path(task_id))

        source_df = task.get("source_df")
        src_path = _get_ai_task_source_df_path(task_id)
        if isinstance(source_df, pd.DataFrame) and (not src_path.exists()):
            source_df.to_pickle(src_path)

        meta = {k: v for k, v in task.items() if k not in ("df_work", "source_df")}
        meta["updated_at"] = now_iso()

        meta_path = _get_ai_task_meta_path(task_id)
        tmp_meta_path = task_dir / "meta.tmp.json"
        with open(tmp_meta_path, "w", encoding="utf-8") as f:
            json.dump(_json_safe(meta), f, ensure_ascii=False, indent=2)
        os.replace(tmp_meta_path, meta_path)


def load_ai_task_state(task_id: str) -> Optional[Dict[str, Any]]:
    if not task_id:
        return None

    meta_path = _get_ai_task_meta_path(task_id)
    if not meta_path.exists():
        return None

    with _AI_TASK_FILE_LOCK:
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
        except Exception:
            return None

        if not isinstance(meta, dict):
            return None

        df_path = _get_ai_task_df_path(task_id)
        src_path = _get_ai_task_source_df_path(task_id)
        try:
            meta["df_work"] = pd.read_pickle(df_path) if df_path.exists() else pd.DataFrame()
        except Exception:
            meta["df_work"] = pd.DataFrame()
        try:
            meta["source_df"] = pd.read_pickle(src_path) if src_path.exists() else pd.DataFrame()
        except Exception:
            meta["source_df"] = pd.DataFrame()
        return meta


def load_latest_ai_task_state(prefer_active: bool = True) -> Optional[Dict[str, Any]]:
    root = get_task_root_path()
    if not root.exists():
        return None

    candidates: List[Tuple[float, str, str]] = []
    for meta_path in root.glob("*/meta.json"):
        task_id = meta_path.parent.name
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
        except Exception:
            continue
        if not isinstance(meta, dict):
            continue

        status = str(meta.get("status", ""))
        if prefer_active and status not in (AI_TASK_STATUS_RUNNING, AI_TASK_STATUS_PAUSED, AI_TASK_STATUS_ERROR):
            continue
        ts = float(meta_path.stat().st_mtime)
        candidates.append((ts, task_id, status))

    if not candidates and prefer_active:
        return load_latest_ai_task_state(prefer_active=False)
    if not candidates:
        return None

    candidates.sort(reverse=True)
    return load_ai_task_state(candidates[0][1])


def split_ai_task_frames(task: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    df_work = task.get("df_work")
    if not isinstance(df_work, pd.DataFrame) or df_work.empty:
        empty = pd.DataFrame()
        return {"ok": empty, "bad": empty, "pending": empty, "processed": empty}

    total = max(0, min(int(task.get("total", len(df_work))), len(df_work)))
    in_scope = df_work.iloc[:total].copy()
    out_scope = df_work.iloc[total:].copy()

    if COL_AI_TASK_STATUS in in_scope.columns:
        processed_mask = in_scope[COL_AI_TASK_STATUS] == AI_ROW_STATUS_DONE
    else:
        if COL_AI_MATCH in in_scope.columns:
            processed_mask = in_scope[COL_AI_MATCH].notna()
        else:
            processed_mask = pd.Series(False, index=in_scope.index)

    processed = in_scope[processed_mask].copy()
    pending_scope = in_scope[~processed_mask].copy()
    pending = pd.concat([pending_scope, out_scope], axis=0).copy()

    ok = processed[processed[COL_AI_MATCH] == True].copy() if COL_AI_MATCH in processed.columns else pd.DataFrame()
    bad = processed[processed[COL_AI_MATCH] != True].copy() if COL_AI_MATCH in processed.columns else pd.DataFrame()

    return {"ok": ok, "bad": bad, "pending": pending, "processed": processed}


def summarize_ai_task(task: Dict[str, Any]) -> Dict[str, Any]:
    frames = split_ai_task_frames(task)
    total = int(task.get("total", 0))
    processed_rows = len(frames["processed"])
    return {
        "total": total,
        "processed_rows": processed_rows,
        "pending_rows": max(0, total - processed_rows),
        "ok_rows": len(frames["ok"]),
        "bad_rows": len(frames["bad"]),
    }


def task_status_label(status: str) -> str:
    m = {
        AI_TASK_STATUS_RUNNING: "运行中",
        AI_TASK_STATUS_PAUSED: "已暂停",
        AI_TASK_STATUS_COMPLETED: "已完成",
        AI_TASK_STATUS_ERROR: "错误",
    }
    return m.get(str(status), str(status))


def _json_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    return str(value)


def append_operation_history(stage: str, action: str, detail: Dict[str, Any]) -> None:
    now = datetime.now()
    record = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "year": now.strftime("%Y"),
        "month": now.strftime("%m"),
        "day": now.strftime("%d"),
        "stage": stage,
        "action": action,
        "operator": os.getenv("USERNAME", ""),
    }
    record.update(_json_safe(detail))
    line = json.dumps(record, ensure_ascii=False)

    try:
        with open(get_history_file_path(), "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception as e:
        st.session_state["history_write_error"] = str(e)


def load_operation_history_df() -> pd.DataFrame:
    path = get_history_file_path()
    if not path.exists():
        return pd.DataFrame()

    records: List[Dict[str, Any]] = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                try:
                    obj = json.loads(s)
                    if isinstance(obj, dict):
                        records.append(obj)
                except Exception:
                    continue
    except Exception as e:
        st.session_state["history_read_error"] = str(e)
        return pd.DataFrame()

    if not records:
        return pd.DataFrame()
    df_hist = pd.DataFrame(records)
    if "timestamp" in df_hist.columns:
        df_hist = df_hist.sort_values("timestamp", ascending=False).reset_index(drop=True)
    return df_hist


def build_history_download_name(prefix: str = "操作历史") -> str:
    return f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{prefix}.csv"


_WIDGET_KEY_CALL_COUNTS: Dict[str, int] = {}


def unique_widget_key(base: str) -> str:
    """
    生成当前脚本运行周期内的唯一 widget key。
    用于部署环境出现代码块重复执行时，避免 StreamlitDuplicateElementKey。
    """
    count = _WIDGET_KEY_CALL_COUNTS.get(base, 0) + 1
    _WIDGET_KEY_CALL_COUNTS[base] = count
    return f"{base}__{count}"


def render_preview_dataframe(
    df: pd.DataFrame,
    *,
    title: str,
    key_prefix: str,
    default_rows: int = 50,
    height: int = 380,
    expanded: bool = False,
) -> None:
    """可折叠的数据预览：按需展开，并支持选择展示条数。"""
    total_rows = int(len(df)) if isinstance(df, pd.DataFrame) else 0
    with st.expander(f"{title}（共 {total_rows} 行）", expanded=expanded):
        if total_rows <= 0:
            st.info("暂无数据。")
            return

        row_options: List[Any] = [20, 50, 100, 200, 500, "全部"]
        if default_rows not in row_options:
            default_rows = 50

        selected_rows = st.selectbox(
            "展示条数",
            options=row_options,
            index=row_options.index(default_rows),
            key=f"{key_prefix}_preview_rows",
            format_func=lambda v: f"{v} 条" if isinstance(v, int) else str(v),
        )

        if selected_rows == "全部":
            df_to_show = df
        else:
            df_to_show = df.head(int(selected_rows))

        st.caption(f"当前显示 {len(df_to_show)} / {total_rows} 行")
        st.dataframe(df_to_show, use_container_width=True, height=height)


def safe_strip_columns(df: pd.DataFrame) -> pd.DataFrame:
    """【强防坑要求】读取后立刻 strip 列名"""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def get_uploaded_bytes(uploaded_file) -> bytes:
    """UploadedFile -> bytes（避免文件指针反复读取问题）"""
    if uploaded_file is None:
        return b""
    try:
        return uploaded_file.getvalue()
    except Exception:
        try:
            uploaded_file.seek(0)
            return uploaded_file.read()
        except Exception:
            return b""


def read_table(uploaded_file) -> pd.DataFrame:
    """读取 xlsx/xls/csv，并在第一时间 strip 列名"""
    if uploaded_file is None:
        return pd.DataFrame()

    filename = uploaded_file.name.lower()
    try:
        if filename.endswith((".xlsx", ".xls")):
            uploaded_file.seek(0)
            # 关键字段（订单号/账号/单号）需保持文本，避免科学计数法和精度风险
            df = pd.read_excel(
                uploaded_file,
                engine="openpyxl",
                dtype=str,
                keep_default_na=False,
                na_filter=False
            )  # 默认第一个sheet
        elif filename.endswith(".csv"):
            uploaded_file.seek(0)
            try:
                df = pd.read_csv(
                    uploaded_file,
                    encoding="utf-8",
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False
                )
            except UnicodeDecodeError:
                uploaded_file.seek(0)
                df = pd.read_csv(
                    uploaded_file,
                    encoding="gbk",
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False
                )
        else:
            raise ValueError("仅支持 .xlsx / .xls / .csv")
    except Exception as e:
        raise RuntimeError(f"文件读取失败：{e}")

    return safe_strip_columns(df)


def find_first_existing_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = set(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def ensure_required_columns(df: pd.DataFrame, required_map: Dict[str, List[str]]) -> Dict[str, str]:
    matched = {}
    missing = []
    cols = set(df.columns)
    for desc, candidates in required_map.items():
        col = None
        for candidate in candidates:
            if candidate in cols:
                col = candidate
                break
        if not col:
            missing.append(f"{desc}（候选：{candidates}）")
        else:
            matched[desc] = col
    if missing:
        raise ValueError("缺少必要列：\n- " + "\n- ".join(missing))
    return matched


@lru_cache(maxsize=8192)
def _parse_money_text(value_text: str) -> Optional[float]:
    s2 = REGEX_MONEY_CLEAN.sub("", value_text)
    if s2 in ("", ".", "-", "-."):
        return None
    try:
        return float(s2)
    except Exception:
        return None


def parse_money(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace("￥", "").replace("¥", "").replace("元", "").replace(",", "").strip()
    return _parse_money_text(s)


@lru_cache(maxsize=16384)
def _normalize_logistics_text(value: str) -> str:
    return REGEX_NON_ALNUM.sub("", value).strip()


def normalize_logistics_no(raw: Any) -> str:
    if raw is None or (isinstance(raw, float) and math.isnan(raw)):
        return ""
    return _normalize_logistics_text(str(raw).strip())


def validate_row(amount: Any, alipay_account: Any, alipay_name: Any, logistics_no: Any) -> Tuple[bool, str]:
    reasons = []

    money = parse_money(amount)
    if money is None:
        reasons.append("金额异常（非数字）")
    else:
        if money > MAX_REFUND_AMOUNT:
            reasons.append("金额异常（金额超标）")

    acct = "" if alipay_account is None else str(alipay_account).strip()
    if acct == "" or (not REGEX_PHONE.match(acct) and not REGEX_EMAIL.match(acct)):
        reasons.append("账号异常（支付宝账号格式不符）")

    name = "" if alipay_name is None else str(alipay_name).strip()
    if name == "" or not REGEX_CN_NAME.match(name):
        reasons.append("实名异常（需2~5个汉字）")

    lno = normalize_logistics_no(logistics_no)
    if lno == "" or not REGEX_LOGISTICS.match(lno):
        reasons.append("单号异常（物流单号需10~16位字母数字且包含数字）")

    if reasons:
        return False, "；".join(reasons)
    return True, ""


# =============================================================================
# 【2.5】入库单号上传与匹配
# =============================================================================

def build_inbound_set(df_inbound: pd.DataFrame, logistics_col: str) -> set:
    """把入库表指定列转为标准化单号集合"""
    normalized_values = (normalize_logistics_no(v) for v in df_inbound[logistics_col].tolist())
    return {v for v in normalized_values if v}


def attach_inbound_flag(df: pd.DataFrame, logistics_col: str, inbound_set: set) -> pd.DataFrame:
    """在 Tab1 的 df 上标记是否已入库"""
    df = df.copy()

    if COL_INBOUND_FLAG not in df.columns:
        df[COL_INBOUND_FLAG] = ""
    if COL_INBOUND_NOTE not in df.columns:
        df[COL_INBOUND_NOTE] = ""

    if not inbound_set:
        return df

    def _flag(x):
        n = normalize_logistics_no(x)
        return "已入库" if n in inbound_set else ""

    df[COL_INBOUND_FLAG] = df[logistics_col].apply(_flag)
    df[COL_INBOUND_NOTE] = df[COL_INBOUND_FLAG].apply(lambda v: "匹配到已入库表" if v == "已入库" else "")
    return df


# =============================================================================
# 【3】链接解析 + Excel 超链接抽取/写回
# =============================================================================

def _dedupe_preserve_order(values: List[str], max_items: Optional[int] = None) -> List[str]:
    seen = set()
    out: List[str] = []
    for v in values:
        item = v.strip()
        if not item or item in seen:
            continue
        out.append(item)
        seen.add(item)
        if max_items is not None and len(out) >= max_items:
            break
    return out


def extract_urls_from_cell(cell_value: Any) -> List[str]:
    """从单元格文本中提取 URL（支持 markdown/裸链）"""
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    urls: List[str] = []
    urls.extend(REGEX_URL_IN_PARENS.findall(s))
    urls.extend(REGEX_URL_GENERIC.findall(s))
    # 去重保持顺序
    return _dedupe_preserve_order(urls)


@lru_cache(maxsize=4096)
def _normalize_preview_url_cached(url: str) -> Tuple[str, ...]:
    if not url:
        return tuple()
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        if "url" in qs:
            raw = unquote(qs["url"][0])
            parts = REGEX_PREVIEW_SPLIT.split(raw)
            extracted = [p.strip() for p in parts if p.strip().startswith("http")]
            if extracted:
                return tuple(extracted)
    except Exception:
        pass
    return (url,)


def normalize_preview_url(url: str) -> List[str]:
    """
    兼容 exportFilePreview?url=... 形式，把真实链接拆出来
    示例：
    https://work.bytenew.com/app.html#/exportFilePreview?url=https%3A%2F%2F...jpeg__https%3A%2F%2F...jpeg
    """
    if not url:
        return []
    return list(_normalize_preview_url_cached(str(url).strip()))


def pick_first_image_url(urls: List[str]) -> Optional[str]:
    """优先挑图片链接，否则返回第一个 http(s)"""
    if not urls:
        return None
    expanded = []
    for u in urls:
        expanded.extend(normalize_preview_url(u))
    for u in expanded:
        if u.lower().endswith(IMAGE_EXTENSIONS):
            return u
    for u in expanded:
        if u.startswith("http"):
            return u
    return None


def pick_image_urls(urls: List[str], max_images: int = 4) -> List[str]:
    """
    ✅ 多图：从单元格里解析出的链接，拆出预览里的多图（__ 分隔），取前 max_images 张
    """
    if not urls:
        return []
    expanded: List[str] = []
    for u in urls:
        expanded.extend(normalize_preview_url(u))

    # 只保留图片直链
    imgs = [u for u in expanded if u.lower().endswith(IMAGE_EXTENSIONS)]

    # 去重保持顺序
    out = _dedupe_preserve_order(imgs, max_items=max_images)

    # 兜底：没有图片后缀时也保留 http（有些链接可能没后缀但可访问）
    if not out:
        http_urls = [u for u in expanded if u.startswith("http")]
        out = _dedupe_preserve_order(http_urls, max_items=max_images)

    return out


@lru_cache(maxsize=8192)
def _extract_image_urls_from_text(raw_text: str, max_images: int) -> Tuple[str, ...]:
    urls = extract_urls_from_cell(raw_text)
    img_urls = pick_image_urls(urls, max_images=max_images)
    if img_urls:
        return tuple(img_urls)

    if raw_text.startswith("http"):
        expanded = normalize_preview_url(raw_text)
        image_candidates = [u for u in expanded if u.lower().endswith(IMAGE_EXTENSIONS)]
        if image_candidates:
            return tuple(image_candidates[:max_images])
        return tuple(expanded[:max_images])
    return tuple()


def extract_image_urls_from_cell_value(cell_value: Any, max_images: int = 4) -> List[str]:
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    raw_text = str(cell_value).strip()
    if not raw_text:
        return []
    return list(_extract_image_urls_from_text(raw_text, int(max_images)))


def extract_hyperlinks_from_excel(file_bytes: bytes, target_header: str, n_rows: Optional[int] = None) -> List[Optional[str]]:
    """
    从 Excel 中提取指定列每行的链接 URL，兼容：
    1) cell.hyperlink.target（原生超链接）
    2) =HYPERLINK("url","预览") / =HYPERLINK("url";"预览") 公式
    3) tooltip / comment 里藏 URL（很多导出“预览”就是这种）
    并且按 n_rows 对齐 pandas 的数据行数（避免 ws.max_row 导致错位）
    """
    if not file_bytes:
        return []
    try:
        # ✅ 关键：data_only=False 才能拿到公式本体
        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        # ✅ 与 pandas 默认 sheet 对齐：第一个工作表
        ws = wb.worksheets[0]

        # 读取表头（第1行）
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            headers.append(str(v).strip() if v is not None else "")

        if target_header not in headers:
            return []

        col_idx = headers.index(target_header) + 1

        # 公式超链接：支持 , 或 ; 分隔
        re_hyper = REGEX_EXCEL_HYPERLINK_FORMULA
        # URL 兜底抽取
        re_url = REGEX_EXCEL_URL_FALLBACK

        # ✅ 行数对齐：只取 pandas 实际行数
        if n_rows is None:
            end_row = ws.max_row
        else:
            end_row = 1 + int(n_rows)  # 表头1行 + n_rows数据行

        links: List[Optional[str]] = []
        for r in range(2, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            url = None

            # 1) 原生 hyperlink
            if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                url = str(cell.hyperlink.target).strip()

            # 2) HYPERLINK公式
            if not url:
                v = cell.value
                if isinstance(v, str):
                    m = re_hyper.search(v)
                    if m:
                        url = m.group(1).strip()

            # 3) tooltip / comment 兜底（黄色提示框常在这里）
            if not url:
                tip = None
                try:
                    tip = getattr(cell.hyperlink, "tooltip", None) if cell.hyperlink else None
                except Exception:
                    tip = None
                if isinstance(tip, str):
                    m = re_url.search(tip)
                    if m:
                        url = m.group(1).strip()

            if not url and cell.comment and isinstance(cell.comment.text, str):
                m = re_url.search(cell.comment.text)
                if m:
                    url = m.group(1).strip()

            links.append(url if url else None)

        return links
    except Exception:
        return []


def attach_hyperlink_helper_column(df: pd.DataFrame, file_bytes: bytes, screenshot_col: str) -> pd.DataFrame:
    """
    若 Excel 截图列是“超链接/公式超链接/tooltip/批注”，pandas 读不到 target，
    用 openpyxl 抽出 URL 存入辅助列 {screenshot_col}__hyperlink
    ✅ 强制按 df 行数对齐，不再要求 len(links)==len(df)
    """
    df = df.copy()
    if not file_bytes or df.empty:
        return df

    links = extract_hyperlinks_from_excel(file_bytes, screenshot_col, n_rows=len(df))

    # 对齐：不足补 None，超出截断
    if len(links) < len(df):
        links = links + [None] * (len(df) - len(links))
    elif len(links) > len(df):
        links = links[:len(df)]

    df[screenshot_col + HYPERLINK_SUFFIX] = links
    return df


def _is_identifier_column(col_name: str) -> bool:
    name = str(col_name)
    return any(k in name for k in IDENTIFIER_COLUMN_KEYWORDS)


def _normalize_scientific_text(s: str) -> str:
    v = s.strip()
    if not v or not REGEX_SCI_NUMBER.match(v):
        return v
    try:
        d = Decimal(v)
    except InvalidOperation:
        return v
    if d == d.to_integral_value():
        return format(d.quantize(Decimal("1")), "f")
    out = format(d, "f")
    out = out.rstrip("0").rstrip(".")
    return out if out else "0"


def _normalize_identifier_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return format(value, ".0f")
        out = format(value, "f").rstrip("0").rstrip(".")
        return out if out else "0"
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s:
        return ""
    return _normalize_scientific_text(s)


def find_column_with_fallback(
    df: pd.DataFrame,
    candidates: List[str],
    fuzzy_keywords: Optional[List[str]] = None
) -> Optional[str]:
    col = find_first_existing_column(df, candidates)
    if col:
        return col
    if not fuzzy_keywords:
        return None

    for c in df.columns:
        name = str(c).strip()
        lname = name.lower()
        if any(k.lower() in lname for k in fuzzy_keywords):
            return c
    return None


def build_row_identity_keys(df: pd.DataFrame, id_col: str, order_col: str, logistics_col: str) -> Tuple[List[Tuple[str, str, str]], List[str]]:
    keys: List[Tuple[str, str, str]] = []
    logistics_keys: List[str] = []

    id_values = df[id_col].tolist()
    order_values = df[order_col].tolist()
    logistics_values = df[logistics_col].tolist()

    for id_v, order_v, logistics_v in zip(id_values, order_values, logistics_values):
        id_key = _normalize_identifier_cell(id_v)
        order_key = _normalize_identifier_cell(order_v)
        logistics_key = normalize_logistics_no(logistics_v)
        keys.append((id_key, order_key, logistics_key))
        if logistics_key:
            logistics_keys.append(logistics_key)

    return keys, logistics_keys


def compare_source_and_processed(
    source_df: pd.DataFrame,
    processed_df: pd.DataFrame,
    stage_name: str
) -> Dict[str, Any]:
    report: Dict[str, Any] = {
        "stage": stage_name,
        "can_compare": False,
        "ok": False,
        "source_rows": len(source_df),
        "processed_rows": len(processed_df),
    }

    if source_df is None or processed_df is None or source_df.empty or processed_df.empty:
        report["message"] = "源数据或处理后数据为空，无法校验。"
        return report

    src_id = find_column_with_fallback(source_df, COL_ID_CANDIDATES, fuzzy_keywords=["id", "旺旺"])
    src_order = find_column_with_fallback(source_df, COL_ORDER_NO_CANDIDATES, fuzzy_keywords=["订单"])
    src_lno = find_column_with_fallback(source_df, COL_LOGISTICS_NO_CANDIDATES, fuzzy_keywords=["物流", "快递"])

    dst_id = find_column_with_fallback(processed_df, COL_ID_CANDIDATES, fuzzy_keywords=["id", "旺旺"])
    dst_order = find_column_with_fallback(processed_df, COL_ORDER_NO_CANDIDATES, fuzzy_keywords=["订单"])
    dst_lno = find_column_with_fallback(processed_df, COL_LOGISTICS_NO_CANDIDATES, fuzzy_keywords=["物流", "快递"])

    missing = []
    if not src_id or not dst_id:
        missing.append("ID")
    if not src_order or not dst_order:
        missing.append("订单号")
    if not src_lno or not dst_lno:
        missing.append("物流单号")
    if missing:
        report["message"] = "缺少对比字段：" + "、".join(missing)
        report["mapping"] = {
            "source_id_col": src_id or "",
            "source_order_col": src_order or "",
            "source_logistics_col": src_lno or "",
            "processed_id_col": dst_id or "",
            "processed_order_col": dst_order or "",
            "processed_logistics_col": dst_lno or "",
        }
        return report

    src_keys, src_logistics = build_row_identity_keys(source_df, src_id, src_order, src_lno)
    dst_keys, dst_logistics = build_row_identity_keys(processed_df, dst_id, dst_order, dst_lno)

    src_counter = Counter(src_keys)
    dst_counter = Counter(dst_keys)
    missing_counter = src_counter - dst_counter
    extra_counter = dst_counter - src_counter

    missing_rows = int(sum(missing_counter.values()))
    extra_rows = int(sum(extra_counter.values()))

    src_logistics_counter = Counter(src_logistics)
    dst_logistics_counter = Counter(dst_logistics)
    src_dup_logistics = int(sum(1 for c in src_logistics_counter.values() if c > 1))
    dst_dup_logistics = int(sum(1 for c in dst_logistics_counter.values() if c > 1))

    diff_preview: List[Dict[str, Any]] = []
    for key, cnt in missing_counter.items():
        diff_preview.append({"差异类型": "源有但处理后无", "ID": key[0], "订单号": key[1], "物流单号": key[2], "数量": cnt})
        if len(diff_preview) >= 20:
            break
    if len(diff_preview) < 20:
        for key, cnt in extra_counter.items():
            diff_preview.append({"差异类型": "处理后有但源无", "ID": key[0], "订单号": key[1], "物流单号": key[2], "数量": cnt})
            if len(diff_preview) >= 20:
                break

    report.update({
        "can_compare": True,
        "ok": (missing_rows == 0 and extra_rows == 0),
        "missing_rows": missing_rows,
        "extra_rows": extra_rows,
        "source_duplicate_logistics": src_dup_logistics,
        "processed_duplicate_logistics": dst_dup_logistics,
        "mapping": {
            "source_id_col": src_id,
            "source_order_col": src_order,
            "source_logistics_col": src_lno,
            "processed_id_col": dst_id,
            "processed_order_col": dst_order,
            "processed_logistics_col": dst_lno,
        },
        "diff_preview": diff_preview,
    })
    return report


def render_alignment_report(report: Dict[str, Any], title: str) -> None:
    st.markdown(f"#### 🔍 {title}")

    if not report.get("can_compare"):
        st.warning(f"⚠️ 未完成一致性校验：{report.get('message', '缺少必要字段。')}")
        mapping = report.get("mapping", {})
        if mapping:
            st.caption(
                "字段识别情况："
                f"源[ID:{mapping.get('source_id_col', '无')} / 订单:{mapping.get('source_order_col', '无')} / 物流:{mapping.get('source_logistics_col', '无')}]；"
                f"结果[ID:{mapping.get('processed_id_col', '无')} / 订单:{mapping.get('processed_order_col', '无')} / 物流:{mapping.get('processed_logistics_col', '无')}]"
            )
        return

    if report.get("ok"):
        st.success("✅ 一致性校验通过：处理后数据与源数据按 ID/订单号/物流单号 完整对应。")
    else:
        st.error("❌ 一致性校验未通过：发现源数据与处理后数据存在不一致。请先确认后再继续。")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("源数据行数", int(report.get("source_rows", 0)))
    c2.metric("处理后行数", int(report.get("processed_rows", 0)))
    c3.metric("源有结果无", int(report.get("missing_rows", 0)))
    c4.metric("结果有源无", int(report.get("extra_rows", 0)))

    mapping = report.get("mapping", {})
    st.caption(
        "字段映射："
        f"源[ID:{mapping.get('source_id_col', '无')} / 订单:{mapping.get('source_order_col', '无')} / 物流:{mapping.get('source_logistics_col', '无')}]；"
        f"结果[ID:{mapping.get('processed_id_col', '无')} / 订单:{mapping.get('processed_order_col', '无')} / 物流:{mapping.get('processed_logistics_col', '无')}]"
    )

    src_dup = int(report.get("source_duplicate_logistics", 0))
    dst_dup = int(report.get("processed_duplicate_logistics", 0))
    if src_dup > 0 or dst_dup > 0:
        st.warning(f"⚠️ 物流单号重复检查：源数据重复 {src_dup} 个，处理后重复 {dst_dup} 个。")

    diff_preview = report.get("diff_preview", [])
    if diff_preview:
        with st.expander("查看差异样例（最多20条）", expanded=False):
            st.dataframe(pd.DataFrame(diff_preview), use_container_width=True, height=260)

def df_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "sheet1",
    hyperlink_cols: Optional[List[str]] = None
) -> bytes:
    """
    DataFrame -> Excel bytes，并为指定列写回超链接：
    - 保留原文字（例如“预览/浏览”/长串文本），但整格可点击
    - 优先使用辅助列 {col}__hyperlink 的 target
    - 若无辅助列，则尝试从 cell.value 文本解析 URL（包含预览链接时也能拆 __）
    """
    df_export = df.copy()
    identifier_cols = [c for c in df_export.columns if _is_identifier_column(c)]
    for col in identifier_cols:
        df_export[col] = df_export[col].map(_normalize_identifier_cell)

    # 记录每个超链接列对应的 target 列表（按行对齐）
    link_targets: Dict[str, List[Optional[str]]] = {}

    if hyperlink_cols:
        for col in hyperlink_cols:
            helper_col = col + HYPERLINK_SUFFIX
            if helper_col in df_export.columns:
                link_targets[col] = df_export[helper_col].tolist()
                # 导出表里不带辅助列
                df_export.drop(columns=[helper_col], inplace=True, errors="ignore")

    # 先写 Excel
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)

    if not hyperlink_cols:
        return bio.getvalue()

    # 再用 openpyxl 写回 hyperlink
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # 表头映射：列名->列号
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            header_map[str(v).strip()] = c

    # 关键标识字段强制文本格式，避免 Excel 科学计数法显示
    for col in identifier_cols:
        if col not in header_map:
            continue
        cidx = header_map[col]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=cidx)
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            cell.number_format = "@"

    for col in hyperlink_cols:
        if col not in header_map:
            continue
        cidx = header_map[col]
        targets = link_targets.get(col, [])

        # Excel 数据行从第2行开始，DataFrame 行从 0 开始
        for r in range(2, ws.max_row + 1):
            df_idx = r - 2
            cell = ws.cell(row=r, column=cidx)

            # 1) 优先取辅助列 target（原 Excel 超链接/公式解析/tooltip解析出来的URL）
            target = None
            if targets and df_idx < len(targets):
                target = targets[df_idx]

            # 2) 若没有 target，则尝试从 cell.value 文本解析（适配长串 url 文本/预览链接）
            if not target:
                val = "" if cell.value is None else str(cell.value).strip()
                urls = extract_urls_from_cell(val)
                # 这里使用多图逻辑，取第一张当 hyperlink（Excel 单格只能挂一个）
                imgs = pick_image_urls(urls, max_images=1)
                target = imgs[0] if imgs else None
                if not target and val.startswith("http"):
                    # 如果是预览链接，尝试拆
                    expanded = normalize_preview_url(val)
                    target = expanded[0] if expanded else val

            if target and str(target).startswith("http"):
                cell.hyperlink = str(target).strip()
                cell.style = "Hyperlink"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# =============================================================================
# 【4】DashScope 调用（Tab2）+ 多图提示词 + 速率限制与重试
# =============================================================================

def get_dashscope_api_key() -> str:
    key = ""
    try:
        key = st.secrets.get("DASHSCOPE_API_KEY", "")
    except Exception:
        key = ""
    if not key:
        key = os.getenv("DASHSCOPE_API_KEY", "")
    return key


def make_vl_prompt(expected_amount: float) -> str:
    """
    ✅ 优化提示词：多图场景下从所有图片中找“运费/快递费/配送费/邮费”对应金额，避免把商品金额当运费
    """
    return f"""
你是电商售后财务审核助手。用户可能上传了多张截图（同一条售后记录的图片从上到下依次排列）。
你的任务是在所有图片中寻找“寄回运费/快递费/配送费/邮费/寄件费用/实付运费/总运费”等字段对应的金额（单位：元），并与用户填写金额进行核对。

用户填写的退回运费金额 expected_amount = {expected_amount:.2f} 元。

【重要规则（避免误判）】
1) 只把与“运费/快递费/配送费/邮费/寄件费用/运费金额/实付运费/总运费”明确相关的金额当作运费。
2) 如果图片里出现“商品金额/订单金额/合计/实付/优惠/退款金额/支付金额”等多个金额：
   - 优先选择紧邻“运费/快递费/配送费/邮费/寄件费用”文字的金额。
   - 不要把商品金额当运费。
3) 运费可能显示为 0、0.00、¥0、免运费，也要识别为 0。
4) 若所有图片都没有明确的“运费/快递费/配送费/邮费/寄件费用”字段或对应金额，请返回 paid_amount = null，并在 reason 写清楚：
   - “多图中未找到运费字段”
   - 或 “图片为商品瑕疵/聊天/订单页，非运费截图”
   - 或 “图片模糊/遮挡无法识别运费金额”
5) 金额比对：允许误差 0.01。相等则 is_match=true，否则 false；识别不到则 is_match=null。
6) 禁止猜测：看不清/不确定就返回 null。

【输出格式】
只输出 JSON（不要任何额外文字），字段必须包含：
- paid_amount: 数字或 null
- is_match: true/false 或 null
- reason: 简短明确说明（如：一致/不一致/多图未找到运费字段/图片模糊无法识别/非运费截图等）
可选字段（若能提供更好）：
- image_index: 找到运费的图片序号（从1开始；未找到则 null）
- evidence_text: 支撑判断的关键词片段（例如“运费 ¥8.00”）
- confidence: 0~1 的置信度（无法判断则 0）
""".strip()


def _looks_like_rate_limited(text: str) -> bool:
    if not text:
        return False
    t = str(text).lower()
    keywords = ["rate", "limit", "throttle", "too many", "busy", "繁忙", "限流", "频率", "qps", "quota", "exceeded", "429"]
    return any(k in t for k in keywords)


def _sleep_with_ui(seconds: float, ui_slot=None):
    if seconds <= 0:
        return
    if ui_slot is not None:
        ui_slot.info(f"⏳ 触发速率限制/退避等待 {seconds:.2f}s …")
    time.sleep(seconds)


def _parse_vl_json(raw_text: str, expected_amount: float) -> Dict[str, Any]:
    """解析模型输出为 dict，并做金额/一致性兜底"""
    data = None
    try:
        data = json.loads(raw_text)
    except Exception:
        m = re.search(r"\{[\s\S]*\}", raw_text)
        if m:
            try:
                data = json.loads(m.group(0))
            except Exception:
                data = None

    if not isinstance(data, dict):
        return {"paid_amount": None, "is_match": None, "reason": "输出非JSON", "raw_text": raw_text}

    paid_f = parse_money(data.get("paid_amount")) if data.get("paid_amount") is not None else None
    is_match = data.get("is_match", None)
    if is_match is None and paid_f is not None:
        is_match = (abs(paid_f - expected_amount) <= 0.01)

    reason = str(data.get("reason", "")).strip()
    if not reason:
        reason = "一致" if is_match else ("图片模糊或无法识别金额" if paid_f is None else "不一致")

    # 把可选字段透传（不影响你现有表结构）
    out = {
        "paid_amount": paid_f,
        "is_match": is_match,
        "reason": reason,
        "raw_text": raw_text
    }
    for k in ["image_index", "evidence_text", "confidence", "candidates"]:
        if k in data:
            out[k] = data.get(k)
    return out


def call_qwen_vl_extract_amount_multi(image_urls: List[str], expected_amount: float, api_key: str, model: str) -> Dict[str, Any]:
    """
    ✅ 多图调用：把同一条记录的多张图片一起发给模型，让模型在多图中找运费金额
    """
    if dashscope is None:
        return {"paid_amount": None, "is_match": None, "reason": "DashScope SDK 未安装（pip install dashscope）", "raw_text": ""}
    if not api_key:
        return {"paid_amount": None, "is_match": None, "reason": "缺少 DASHSCOPE_API_KEY", "raw_text": ""}
    if not image_urls:
        return {"paid_amount": None, "is_match": None, "reason": "图片URL为空", "raw_text": ""}

    prompt = make_vl_prompt(expected_amount)

    # 多图：先放图片，再放文本 prompt
    content = [{"image": u} for u in image_urls]
    content.append({"text": prompt})
    messages = [{"role": "user", "content": content}]

    try:
        resp = dashscope.MultiModalConversation.call(api_key=api_key, model=model, messages=messages)

        if hasattr(resp, "status_code") and HTTPStatus is not None and resp.status_code != HTTPStatus.OK:
            msg = f"API失败：{getattr(resp, 'code', '')} {getattr(resp, 'message', '')}".strip()
            return {
                "paid_amount": None,
                "is_match": None,
                "reason": msg,
                "raw_text": str(resp),
                "status_code": getattr(resp, "status_code", None)
            }

        raw_text = ""
        try:
            raw_text = resp.output.choices[0]["message"]["content"][0]["text"]
        except Exception:
            raw_text = str(resp)

        return _parse_vl_json(raw_text, expected_amount)

    except Exception as e:
        return {"paid_amount": None, "is_match": None, "reason": f"异常：{e}", "raw_text": traceback.format_exc()}


def call_qwen_vl_extract_amount_multi_with_rl(
    image_urls: List[str],
    expected_amount: float,
    api_key: str,
    model: str,
    min_interval_sec: float = 0.8,
    max_retries: int = 4,
    backoff_base_sec: float = 1.0,
    ui_slot=None,
) -> Dict[str, Any]:
    """
    ✅ 多图版本：速率限制 + 自动重试（指数退避）
    """
    if "ai_last_call_ts" not in st.session_state:
        st.session_state["ai_last_call_ts"] = 0.0

    attempt = 0
    while True:
        # 节流：确保请求间隔
        now = time.monotonic()
        last = float(st.session_state.get("ai_last_call_ts", 0.0))
        wait = max(0.0, float(min_interval_sec) - (now - last))
        if wait > 0:
            _sleep_with_ui(wait, ui_slot)

        st.session_state["ai_last_call_ts"] = time.monotonic()
        res = call_qwen_vl_extract_amount_multi(image_urls, expected_amount, api_key, model)

        reason = str(res.get("reason", "") or "")
        status_code = res.get("status_code", None)

        need_retry = False
        if status_code in (429, 503, 502, 504):
            need_retry = True
        if _looks_like_rate_limited(reason):
            need_retry = True
        if reason == "输出非JSON":
            need_retry = True

        if (not need_retry) or (attempt >= int(max_retries)):
            return res

        sleep_s = float(backoff_base_sec) * (2 ** attempt)
        sleep_s = min(sleep_s, 15.0)
        attempt += 1
        _sleep_with_ui(sleep_s, ui_slot)


def _call_qwen_vl_extract_amount_multi_with_rl_worker(
    image_urls: List[str],
    expected_amount: float,
    api_key: str,
    model: str,
    last_call_ts: float,
    min_interval_sec: float = 0.8,
    max_retries: int = 4,
    backoff_base_sec: float = 1.0,
) -> Tuple[Dict[str, Any], float]:
    """
    后台线程版：不依赖 Streamlit session_state。
    返回 (结果, 更新后的 last_call_ts)。
    """
    attempt = 0
    while True:
        now_m = time.monotonic()
        wait = max(0.0, float(min_interval_sec) - (now_m - float(last_call_ts)))
        if wait > 0:
            time.sleep(wait)

        last_call_ts = time.monotonic()
        res = call_qwen_vl_extract_amount_multi(image_urls, expected_amount, api_key, model)

        reason = str(res.get("reason", "") or "")
        status_code = res.get("status_code", None)

        need_retry = False
        if status_code in (429, 503, 502, 504):
            need_retry = True
        if _looks_like_rate_limited(reason):
            need_retry = True
        if reason == "输出非JSON":
            need_retry = True

        if (not need_retry) or (attempt >= int(max_retries)):
            return res, last_call_ts

        sleep_s = float(backoff_base_sec) * (2 ** attempt)
        sleep_s = min(sleep_s, 15.0)
        attempt += 1
        time.sleep(sleep_s)


def _process_ai_task_one_row(task: Dict[str, Any], api_key: str, worker_token: str, last_call_ts: float) -> float:
    df_work = task.get("df_work")
    if not isinstance(df_work, pd.DataFrame) or df_work.empty:
        raise RuntimeError("任务数据为空，无法处理。")

    idx = int(task.get("next_idx", 0))
    total = int(task.get("total", len(df_work)))
    if idx >= total:
        return last_call_ts

    col_amount = str(task.get("col_amount", ""))
    col_shot = str(task.get("col_shot", ""))
    model_name = str(task.get("model_name", DEFAULT_VL_MODEL))
    max_images = int(task.get("max_images", 4))
    min_interval_sec = float(task.get("min_interval_sec", 0.8))
    max_retries = int(task.get("max_retries", 4))
    backoff_base_sec = float(task.get("backoff_base_sec", 1.0))

    row = df_work.iloc[idx]
    expected = parse_money(row.get(col_amount))

    paid_amount = None
    is_match = False
    note = ""

    if expected is None:
        note = "金额字段无法解析为数字（请先回到步骤一检查/修正）"
    else:
        raw_cell = row.get(col_shot + HYPERLINK_SUFFIX) or row.get(col_shot)
        img_urls = extract_image_urls_from_cell_value(raw_cell, max_images=max_images)

        # 兜底：单元格原值是 http 但未被正则提取到。
        if not img_urls and isinstance(raw_cell, str) and raw_cell.strip().startswith("http"):
            expanded = normalize_preview_url(raw_cell.strip())
            img_urls = [u for u in expanded if u.lower().endswith(IMAGE_EXTENSIONS)]
            img_urls = img_urls[:max_images] if img_urls else expanded[:max_images]

        if not img_urls:
            res = {
                "paid_amount": None,
                "is_match": None,
                "reason": "未找到可用图片URL（预览链接未解析出图片）",
                "raw_text": "",
            }
        else:
            res, last_call_ts = _call_qwen_vl_extract_amount_multi_with_rl_worker(
                img_urls,
                float(expected),
                api_key=api_key,
                model=model_name,
                last_call_ts=last_call_ts,
                min_interval_sec=min_interval_sec,
                max_retries=max_retries,
                backoff_base_sec=backoff_base_sec,
            )

        paid_amount = res.get("paid_amount")
        is_match = (res.get("is_match") is True)
        note = "" if is_match else (res.get("reason") or "AI判定异常")

    df_work.at[idx, COL_AI_EXTRACTED_AMOUNT] = paid_amount
    df_work.at[idx, COL_AI_MATCH] = bool(is_match)
    df_work.at[idx, COL_AI_NOTE] = note
    df_work.at[idx, COL_AI_TASK_STATUS] = AI_ROW_STATUS_DONE
    df_work.at[idx, COL_AI_TASK_UPDATED_AT] = now_iso()

    task["df_work"] = df_work
    task["next_idx"] = idx + 1
    task["error_message"] = ""
    task["worker_token"] = worker_token
    if int(task["next_idx"]) >= int(task.get("total", len(df_work))):
        task["status"] = AI_TASK_STATUS_COMPLETED
        task["finished_at"] = now_iso()
    save_ai_task_state(task)
    return last_call_ts


def ai_task_worker_loop(task_id: str, api_key: str, worker_token: str) -> None:
    last_call_ts = 0.0
    while True:
        task = load_ai_task_state(task_id)
        if not task:
            return

        if str(task.get("worker_token", "")) != str(worker_token):
            return

        status = str(task.get("status", ""))
        if status != AI_TASK_STATUS_RUNNING:
            return

        try:
            total = int(task.get("total", 0))
            next_idx = int(task.get("next_idx", 0))
            if next_idx >= total:
                task["status"] = AI_TASK_STATUS_COMPLETED
                task["finished_at"] = now_iso()
                save_ai_task_state(task)
                return

            last_call_ts = _process_ai_task_one_row(task, api_key=api_key, worker_token=worker_token, last_call_ts=last_call_ts)
            time.sleep(0.01)

        except Exception as e:
            task["status"] = AI_TASK_STATUS_ERROR
            task["error_message"] = str(e)
            task["finished_at"] = now_iso()
            save_ai_task_state(task)
            return


def start_ai_task_worker(task_id: str, api_key: str) -> bool:
    if not task_id or not api_key:
        return False

    task = load_ai_task_state(task_id)
    if not task:
        return False

    worker_token = f"w_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    task["worker_token"] = worker_token
    task["status"] = AI_TASK_STATUS_RUNNING
    task["error_message"] = ""
    save_ai_task_state(task)

    t = threading.Thread(
        target=ai_task_worker_loop,
        args=(task_id, api_key, worker_token),
        daemon=True,
        name=f"ai_task_{task_id[-8:]}",
    )
    t.start()
    return True


def finalize_ai_task_if_needed(task: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(task, dict):
        return task
    if str(task.get("status", "")) != AI_TASK_STATUS_COMPLETED:
        return task
    if bool(task.get("history_logged", False)):
        return task

    frames = split_ai_task_frames(task)
    df_ok = frames["ok"]
    df_bad = frames["bad"]
    df_pending = frames["pending"]

    col_shot = str(task.get("col_shot", ""))
    hyperlink_ok = [col_shot] if col_shot and col_shot in df_ok.columns else None
    hyperlink_bad = [col_shot] if col_shot and col_shot in df_bad.columns else None
    hyperlink_pending = [col_shot] if col_shot and col_shot in df_pending.columns else None

    ts = now_ts()
    ok_name = f"{ts}_AI复核正常_待打款.xlsx"
    bad_name = f"{ts}_AI复核异常_需人工.xlsx"
    pending_name = f"{ts}_AI未处理_待继续.xlsx"

    b_ok = df_to_excel_bytes(df_ok, sheet_name="AI复核正常", hyperlink_cols=hyperlink_ok)
    b_bad = df_to_excel_bytes(df_bad, sheet_name="AI复核异常", hyperlink_cols=hyperlink_bad)
    b_pending = df_to_excel_bytes(df_pending, sheet_name="AI未处理", hyperlink_cols=hyperlink_pending) if not df_pending.empty else b""

    ok_artifact = save_artifact_bytes("step3_ai_ok", ok_name, b_ok)
    bad_artifact = save_artifact_bytes("step3_ai_bad", bad_name, b_bad)
    pending_artifact = save_artifact_bytes("step3_ai_pending", pending_name, b_pending) if b_pending else ""

    source_df = task.get("source_df")
    if not isinstance(source_df, pd.DataFrame) or source_df.empty:
        source_df = task.get("df_work", pd.DataFrame()).copy()
    report_step3 = compare_source_and_processed(source_df, task.get("df_work", pd.DataFrame()), stage_name="步骤三AI复核")

    append_operation_history(
        stage="步骤三AI复核",
        action="AI任务完成",
        detail={
            "task_id": task.get("task_id", ""),
            "source_file": task.get("source_file", ""),
            "input_rows": int(task.get("input_rows", 0)),
            "processed_rows": int(task.get("next_idx", 0)),
            "output_rows": len(task.get("df_work", pd.DataFrame())),
            "ai_ok_rows": len(df_ok),
            "ai_bad_rows": len(df_bad),
            "ai_pending_rows": len(df_pending),
            "alignment_can_compare": report_step3.get("can_compare"),
            "alignment_ok": report_step3.get("ok") if report_step3.get("can_compare") else None,
            "alignment_missing_rows": report_step3.get("missing_rows"),
            "alignment_extra_rows": report_step3.get("extra_rows"),
            "artifacts": [p for p in [ok_artifact, bad_artifact, pending_artifact] if p],
        }
    )

    task["history_logged"] = True
    task["alignment_report"] = report_step3
    task["artifacts"] = [p for p in [ok_artifact, bad_artifact, pending_artifact] if p]
    save_ai_task_state(task)
    return load_ai_task_state(str(task.get("task_id", ""))) or task


# =============================================================================
# 【5】Streamlit 页面
# =============================================================================

# 防止部署时脚本被重复拼接/执行，导致标题与控件重复渲染
if globals().get("_REFUND_APP_RENDERED_ONCE", False):
    st.stop()
globals()["_REFUND_APP_RENDERED_ONCE"] = True

st.set_page_config(page_title="退运费智能审核系统｜叠纸心意旗舰店", layout="wide")
st.title("🧾 退运费智能审核系统（内部提效）")
st.caption("Streamlit + Pandas + 通义千问-VL（DashScope）")

st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;700;900&display=swap');

:root {
  --ui-primary: #0f5ea8;
  --ui-secondary: #0f766e;
  --ui-accent: #ea580c;
  --ui-bg: #f5f7fb;
  --ui-card: #ffffff;
  --ui-border: #d8e1ee;
  --ui-text: #0f172a;
  --ui-muted: #475569;
}

html, body, [class*="css"] {
  font-family: "Noto Sans SC", "PingFang SC", "Microsoft YaHei", sans-serif;
}

.stApp {
  background:
    radial-gradient(1100px 360px at -8% -12%, rgba(15, 94, 168, 0.17), transparent 56%),
    radial-gradient(900px 280px at 108% -8%, rgba(15, 118, 110, 0.13), transparent 58%),
    var(--ui-bg);
}

.block-container {
  max-width: 1360px;
  padding-top: 1.2rem;
}

.hero-panel {
  background: linear-gradient(120deg, rgba(15, 94, 168, 0.95), rgba(15, 118, 110, 0.92));
  color: #fff;
  border-radius: 18px;
  padding: 20px 24px;
  margin: 0.2rem 0 1rem 0;
  box-shadow: 0 14px 40px rgba(15, 94, 168, 0.24);
}

.hero-title {
  font-size: 1.32rem;
  font-weight: 900;
  letter-spacing: 0.2px;
  margin-bottom: 0.2rem;
}

.hero-sub {
  font-size: 0.94rem;
  opacity: 0.95;
}

div[data-testid="stTabs"] button {
  border-radius: 999px;
  border: 1px solid var(--ui-border);
  padding: 0.4rem 0.95rem;
  margin-right: 0.32rem;
  color: var(--ui-text);
}

div[data-testid="stTabs"] button[aria-selected="true"] {
  background: linear-gradient(120deg, var(--ui-primary), #1870c0);
  color: #fff;
  border-color: transparent;
  box-shadow: 0 8px 22px rgba(15, 94, 168, 0.28);
}

div[data-testid="stMetric"] {
  background: var(--ui-card);
  border: 1px solid var(--ui-border);
  border-radius: 14px;
  padding: 12px 14px;
  box-shadow: 0 6px 20px rgba(15, 23, 42, 0.06);
}

.ops-flow {
  display: grid;
  grid-template-columns: repeat(3, minmax(0, 1fr));
  gap: 10px;
  margin-top: 0.3rem;
}

.ops-item {
  background: var(--ui-card);
  border: 1px solid var(--ui-border);
  border-radius: 14px;
  padding: 12px 14px;
}

.ops-kicker {
  color: var(--ui-primary);
  font-weight: 800;
  font-size: 0.78rem;
}

.ops-title {
  color: var(--ui-text);
  font-weight: 800;
  margin-top: 4px;
  margin-bottom: 4px;
}

.ops-desc {
  color: var(--ui-muted);
  font-size: 0.86rem;
  line-height: 1.45;
}

.stButton > button, .stDownloadButton > button {
  border-radius: 12px;
  border: 1px solid #c7d4e8;
  font-weight: 700;
}
</style>
    """,
    unsafe_allow_html=True
)

st.markdown(
    """
<div class="hero-panel">
  <div class="hero-title">退运费智能审核中台</div>
  <div class="hero-sub">围绕「清洗 → 入库匹配 → AI复核」构建闭环审核流，支持后台任务、可追溯历史与分段下载。</div>
</div>
    """,
    unsafe_allow_html=True
)

main_tabs = st.tabs([
    "🏠 总览看板",
    "1️⃣ 步骤一：数据清洗",
    "2️⃣ 步骤二&三：入库 + AI复核",
    "🗂️ 历史中心",
])

with main_tabs[0]:
    step1_normal_count = len(st.session_state.get("tab1_normal_df", pd.DataFrame())) if isinstance(st.session_state.get("tab1_normal_df"), pd.DataFrame) else 0
    step1_abnormal_count = len(st.session_state.get("tab1_abnormal_df", pd.DataFrame())) if isinstance(st.session_state.get("tab1_abnormal_df"), pd.DataFrame) else 0
    step2_inbound_count = len(st.session_state.get("step2_inbound_df", pd.DataFrame())) if isinstance(st.session_state.get("step2_inbound_df"), pd.DataFrame) else 0
    step2_pending_count = len(st.session_state.get("step2_not_inbound_df", pd.DataFrame())) if isinstance(st.session_state.get("step2_not_inbound_df"), pd.DataFrame) else 0

    active_task = None
    active_task_id_from_state = str(st.session_state.get("active_ai_task_id", "")).strip()
    if active_task_id_from_state:
        active_task = load_ai_task_state(active_task_id_from_state)
    if active_task is None:
        active_task = load_latest_ai_task_state(prefer_active=True)

    ai_processed_rows = 0
    ai_pending_rows = 0
    ai_task_status = "暂无任务"
    if isinstance(active_task, dict):
        ai_summary = summarize_ai_task(active_task)
        ai_processed_rows = int(ai_summary.get("processed_rows", 0))
        ai_pending_rows = int(ai_summary.get("pending_rows", 0))
        ai_task_status = task_status_label(str(active_task.get("status", "")))

    history_count = len(load_operation_history_df())
    artifact_count = len(load_artifact_catalog_df())

    o1, o2, o3, o4, o5, o6 = st.columns(6)
    o1.metric("清洗正常", step1_normal_count)
    o2.metric("清洗异常", step1_abnormal_count)
    o3.metric("已入库待AI", step2_inbound_count)
    o4.metric("未入库待跟进", step2_pending_count)
    o5.metric("AI已处理", ai_processed_rows)
    o6.metric("AI未处理", ai_pending_rows)
    st.caption(f"当前AI任务状态：{ai_task_status}｜历史记录：{history_count} 条｜历史表格：{artifact_count} 份")

    st.markdown(
        """
<div class="ops-flow">
  <div class="ops-item">
    <div class="ops-kicker">STEP 01</div>
    <div class="ops-title">基础清洗与规则初筛</div>
    <div class="ops-desc">先清洗班牛登记表，按金额、账号、实名、物流单号做规则校验，优先处理异常回访。</div>
  </div>
  <div class="ops-item">
    <div class="ops-kicker">STEP 02</div>
    <div class="ops-title">入库匹配与推进分流</div>
    <div class="ops-desc">将正常数据与已入库单号匹配，生成「可进AI」与「待跟进」两条并行处理路径。</div>
  </div>
  <div class="ops-item">
    <div class="ops-kicker">STEP 03</div>
    <div class="ops-title">AI后台复核与结果落档</div>
    <div class="ops-desc">AI任务支持后台运行、暂停/继续，实时看进度，随时下载已处理与未处理分段结果。</div>
  </div>
</div>
        """,
        unsafe_allow_html=True
    )

    with st.expander("📌 使用说明（展开查看）", expanded=False):
        st.markdown(
            """
- **步骤一（清洗）**：上传班牛登记表 → 清洗 & 规则初筛 → 下载【正常/异常】
  - ✅ 先处理异常回访，再推进后续流程
- **步骤二（入库匹配）**：上传【步骤一正常表/回访后正常表】+【已入库物流单号表】→ 匹配入库状态 → 下载【已入库待AI/未入库待跟进】
- **步骤三（AI复核，多图）**：优先使用步骤二的【已入库待AI】→ 从“预览链接”拆出多张图片 → AI核对运费金额 → 下载【AI正常/AI异常】
  - ✅ 支持后台任务：可暂停/继续，刷新后自动恢复任务状态，并可随时下载【已处理正常/已处理异常/未处理】分段结果

✅ 超链接修复：读取时抽取 URL（兼容 hyperlink / 公式 / tooltip / 批注），导出时写回 hyperlink，不改单元格文字。  
✅ AI 限速：支持最小请求间隔 + 限流/繁忙自动退避重试。  
✅ 多图 AI：支持预览链接里多张图片，模型在多图中自动寻找运费字段对应金额，降低“未显示运费金额”误判。
            """
        )

with main_tabs[3]:
    st.subheader("操作历史与存档中心")
    st.caption("按年月日筛选操作日志，支持查看并下载历史过程产物。")
    if st.session_state.get("history_write_error"):
        st.warning(f"⚠️ 历史记录写入异常：{st.session_state.get('history_write_error')}")
    if st.session_state.get("history_read_error"):
        st.warning(f"⚠️ 历史记录读取异常：{st.session_state.get('history_read_error')}")

    hist_df = load_operation_history_df()
    if hist_df.empty:
        st.info("暂无历史记录。执行步骤一/步骤二/步骤三后会自动记录。")
    else:
        for c in ["year", "month", "day"]:
            if c in hist_df.columns:
                hist_df[c] = hist_df[c].astype(str)
                hist_df[c] = hist_df[c].replace({"nan": "", "None": ""})

        year_options = ["全部"] + sorted([x for x in hist_df["year"].dropna().unique().tolist() if str(x).strip() != ""], reverse=True) if "year" in hist_df.columns else ["全部"]
        sel_year = st.selectbox("年份", options=year_options, index=0, key=unique_widget_key("history_main_filter_year"))

        filtered_df = hist_df.copy()
        if sel_year != "全部" and "year" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["year"] == sel_year]

        month_options = ["全部"] + sorted([x for x in filtered_df["month"].dropna().unique().tolist() if str(x).strip() != ""], reverse=True) if "month" in filtered_df.columns else ["全部"]
        sel_month = st.selectbox("月份", options=month_options, index=0, key=unique_widget_key("history_main_filter_month"))
        if sel_month != "全部" and "month" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["month"] == sel_month]

        day_options = ["全部"] + sorted([x for x in filtered_df["day"].dropna().unique().tolist() if str(x).strip() != ""], reverse=True) if "day" in filtered_df.columns else ["全部"]
        sel_day = st.selectbox("日期", options=day_options, index=0, key=unique_widget_key("history_main_filter_day"))
        if sel_day != "全部" and "day" in filtered_df.columns:
            filtered_df = filtered_df[filtered_df["day"] == sel_day]

        st.caption(f"当前筛选结果：{len(filtered_df)} 条记录。")
        show_cols = [c for c in [
            "timestamp", "stage", "action", "task_id", "operator", "source_file", "input_rows",
            "output_rows", "normal_rows", "abnormal_rows", "inbound_rows", "pending_rows",
            "ai_ok_rows", "ai_bad_rows", "ai_pending_rows", "alignment_ok", "alignment_missing_rows", "alignment_extra_rows", "artifacts"
        ] if c in filtered_df.columns]
        if not show_cols:
            show_cols = filtered_df.columns.tolist()

        st.dataframe(filtered_df[show_cols], use_container_width=True, height=340)

        csv_bytes = filtered_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        excel_bytes = df_to_excel_bytes(filtered_df, sheet_name="操作历史", hyperlink_cols=None)

        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "⬇️ 下载筛选历史（CSV）",
                data=csv_bytes,
                file_name=build_history_download_name("操作历史_筛选结果"),
                mime="text/csv"
            )
        with d2:
            st.download_button(
                "⬇️ 下载筛选历史（Excel）",
                data=excel_bytes,
                file_name=build_history_download_name("操作历史_筛选结果").replace(".csv", ".xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.divider()
        st.markdown("#### 📦 历史表格文件（可下载）")
        artifact_df = load_artifact_catalog_df()
        if artifact_df.empty:
            st.info("暂无已存档表格。完成步骤一/步骤二/步骤三后会自动存档。")
        else:
            for c in ["year", "month", "day"]:
                if c in artifact_df.columns:
                    artifact_df[c] = artifact_df[c].astype(str)

            filtered_artifacts = artifact_df.copy()
            if sel_year != "全部":
                filtered_artifacts = filtered_artifacts[filtered_artifacts["year"] == sel_year]
            if sel_month != "全部":
                filtered_artifacts = filtered_artifacts[filtered_artifacts["month"] == sel_month]
            if sel_day != "全部":
                filtered_artifacts = filtered_artifacts[filtered_artifacts["day"] == sel_day]

            st.caption(f"当前筛选命中的历史表格：{len(filtered_artifacts)} 份。")
            st.dataframe(
                filtered_artifacts[["timestamp", "stage_key", "file_name", "size_kb", "file_path"]],
                use_container_width=True,
                height=260
            )

            if not filtered_artifacts.empty:
                file_options = filtered_artifacts["file_path"].tolist()
                selected_path = st.selectbox("选择要下载的历史表格", options=file_options, key=unique_widget_key("history_artifact_select_path"))
                selected_abs = Path.cwd() / selected_path
                selected_display_name = filtered_artifacts.loc[
                    filtered_artifacts["file_path"] == selected_path, "file_name"
                ].iloc[0]

                if selected_abs.exists():
                    st.download_button(
                        "⬇️ 下载所选历史表格",
                        data=selected_abs.read_bytes(),
                        file_name=selected_display_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                zip_buf = BytesIO()
                with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for _, row in filtered_artifacts.iterrows():
                        rel = str(row.get("file_path", ""))
                        p = Path.cwd() / rel
                        if p.exists():
                            arcname = f"{row.get('timestamp', '').replace(':', '').replace(' ', '_')}__{row.get('file_name', p.name)}"
                            zf.write(p, arcname=arcname)
                st.download_button(
                    "⬇️ 打包下载筛选历史表格（ZIP）",
                    data=zip_buf.getvalue(),
                    file_name=f"{now_ts()}_历史表格.zip",
                    mime="application/zip"
                )

# Sidebar：AI 配置
st.sidebar.header("🔐 AI 配置")
default_key = get_dashscope_api_key()
api_key_input = st.sidebar.text_input("DashScope API Key（secrets/env优先）", value=default_key or "", type="password")
model_name = st.sidebar.text_input("VL 模型", value=DEFAULT_VL_MODEL)

st.sidebar.subheader("🖼️ 多图识别配置")
max_images_per_row = st.sidebar.number_input("每条记录最多传给 AI 的图片数", min_value=1, max_value=10, value=4, step=1)

st.sidebar.subheader("⏱️ AI 速率限制（建议开启）")
min_interval_sec = st.sidebar.number_input("每次请求最小间隔（秒）", min_value=0.0, max_value=10.0, value=0.8, step=0.1)
max_retries = st.sidebar.number_input("限流/繁忙时最大重试次数", min_value=0, max_value=10, value=4, step=1)
backoff_base_sec = st.sidebar.number_input("退避基准秒数（指数退避）", min_value=0.1, max_value=5.0, value=1.0, step=0.1)

max_ai_rows = st.sidebar.number_input("AI 最大处理行数（防误点）", min_value=1, max_value=5000, value=300, step=50)


# =============================================================================
# Tab2：步骤二入库匹配 + 步骤三AI复核（多图）
# =============================================================================
with main_tabs[2]:
    st.subheader("步骤二（A）：入库匹配（先清洗，再匹配）")
    st.info("说明：请先完成步骤一清洗与异常回访，再在这里把【正常表】和【已入库物流单号表】进行匹配。")

    source_left, source_right = st.columns([2, 1])
    with source_left:
        uploaded_step2_source = st.file_uploader(
            "上传【步骤一正常表/回访后正常表】（.xlsx / .xls / .csv）",
            type=["xlsx", "xls", "csv"],
            key="step2_source_uploader"
        )
    with source_right:
        use_tab1_normal_for_step2 = st.checkbox(
            "直接使用步骤一的【正常表】（若已生成）",
            value=True,
            key="use_tab1_normal_for_step2"
        )

    df_step2_source = None
    step2_source_bytes = b""

    if use_tab1_normal_for_step2 and isinstance(st.session_state.get("tab1_normal_df"), pd.DataFrame):
        df_step2_source = st.session_state["tab1_normal_df"].copy()
        st.success("已载入步骤一正常表（当前会话）。")
        render_preview_dataframe(
            df_step2_source,
            title="步骤二待匹配数据预览（来自步骤一）",
            key_prefix="tab2_step1_session_source",
            default_rows=50,
            height=340,
            expanded=False,
        )
    elif uploaded_step2_source is not None:
        try:
            step2_source_bytes = get_uploaded_bytes(uploaded_step2_source)
            df_step2_source = read_table(uploaded_step2_source)

            shot_col_step2_source = find_first_existing_column(df_step2_source, COL_SCREENSHOT_CANDIDATES)
            if shot_col_step2_source and uploaded_step2_source.name.lower().endswith((".xlsx", ".xls")):
                df_step2_source = attach_hyperlink_helper_column(df_step2_source, step2_source_bytes, shot_col_step2_source)

            render_preview_dataframe(
                df_step2_source,
                title="步骤二待匹配数据预览",
                key_prefix="tab2_uploaded_source",
                default_rows=50,
                height=340,
                expanded=False,
            )
        except Exception as e:
            st.error(f"❌ 步骤二待匹配数据读取失败：{e}")
            df_step2_source = None

    inbound_uploader = st.file_uploader(
        "上传【已入库物流单号表】（.xlsx / .xls / .csv）",
        type=["xlsx", "xls", "csv"],
        key="inbound_uploader"
    )

    inbound_set = st.session_state.get("inbound_logistics_set", set())
    inbound_col_name = st.session_state.get("inbound_logistics_col_name", "")

    if inbound_uploader is not None:
        try:
            df_inbound = read_table(inbound_uploader)
            if df_inbound.empty:
                st.warning("读取到的入库表为空，请检查文件内容。")
            else:
                render_preview_dataframe(
                    df_inbound,
                    title="已读取入库表预览",
                    key_prefix="tab2_uploaded_inbound",
                    default_rows=50,
                    height=320,
                    expanded=False,
                )

                required_inb = {"已入库物流单号": COL_LOGISTICS_NO_CANDIDATES}
                matched_inb = ensure_required_columns(df_inbound, required_inb)
                inb_col = matched_inb["已入库物流单号"]

                inbound_set = build_inbound_set(df_inbound, inb_col)
                st.session_state["inbound_logistics_set"] = inbound_set
                st.session_state["inbound_logistics_col_name"] = inb_col
                inbound_col_name = inb_col

                st.success(f"✅ 已缓存入库单号 {len(inbound_set)} 个（列：{inb_col}）。")

        except ValueError as ve:
            st.error(f"❌ 入库表缺少必要列：\n\n{ve}")
        except Exception as e:
            st.error(f"❌ 入库表处理失败：{e}")
            with st.expander("查看错误详情（开发用）"):
                st.code(traceback.format_exc())

    if inbound_set:
        st.caption(f"当前会话已缓存入库单号：{len(inbound_set)} 个（来自列：{inbound_col_name or '未知'}）")
    else:
        st.caption("当前会话尚未缓存入库单号。")

    if use_tab1_normal_for_step2 and isinstance(st.session_state.get("tab1_normal_df"), pd.DataFrame):
        step2_source_label = "步骤一正常表（会话）"
    elif uploaded_step2_source is not None:
        step2_source_label = uploaded_step2_source.name
    else:
        step2_source_label = ""
    inbound_source_label = inbound_uploader.name if inbound_uploader is not None else ("已缓存入库单号（会话）" if inbound_set else "")

    run_match = st.button("🔎 执行入库匹配", type="primary", key="run_inbound_match")

    if run_match:
        if df_step2_source is None or df_step2_source.empty:
            st.warning("请先提供步骤二待匹配数据（步骤一正常表或上传回访后正常表）。")
        elif not inbound_set:
            st.warning("请先上传【已入库物流单号表】。")
        else:
            try:
                required_step2 = {"退回物流单号": COL_LOGISTICS_NO_CANDIDATES}
                matched_step2 = ensure_required_columns(df_step2_source, required_step2)
                col_lno_step2 = matched_step2["退回物流单号"]

                df_step2_matched = attach_inbound_flag(df_step2_source, col_lno_step2, inbound_set)
                df_step2_inbound = df_step2_matched[df_step2_matched[COL_INBOUND_FLAG] == "已入库"].copy()
                df_step2_pending = df_step2_matched[df_step2_matched[COL_INBOUND_FLAG] != "已入库"].copy()
                report_step2 = compare_source_and_processed(df_step2_source, df_step2_matched, stage_name="步骤二入库匹配")

                shot_col_step2 = find_first_existing_column(df_step2_matched, COL_SCREENSHOT_CANDIDATES)
                hyperlink_cols_inbound = [shot_col_step2] if shot_col_step2 and shot_col_step2 in df_step2_inbound.columns else None
                hyperlink_cols_pending = [shot_col_step2] if shot_col_step2 and shot_col_step2 in df_step2_pending.columns else None

                ts_step2 = now_ts()
                inbound_name = f"{ts_step2}_入库匹配通过_待AI复核.xlsx"
                pending_name = f"{ts_step2}_未入库待跟进.xlsx"
                b_inbound = df_to_excel_bytes(df_step2_inbound, sheet_name="已入库", hyperlink_cols=hyperlink_cols_inbound)
                b_pending = df_to_excel_bytes(df_step2_pending, sheet_name="未入库", hyperlink_cols=hyperlink_cols_pending)

                inbound_artifact = save_artifact_bytes("step2_inbound", inbound_name, b_inbound)
                pending_artifact = save_artifact_bytes("step2_pending", pending_name, b_pending)

                st.session_state["step2_matched_df"] = df_step2_matched
                st.session_state["step2_inbound_df"] = df_step2_inbound
                st.session_state["step2_not_inbound_df"] = df_step2_pending
                st.session_state["step2_logistics_col"] = col_lno_step2
                st.session_state["step2_alignment_report"] = report_step2
                st.session_state["step2_inbound_bytes"] = b_inbound
                st.session_state["step2_pending_bytes"] = b_pending
                st.session_state["step2_inbound_name"] = inbound_name
                st.session_state["step2_pending_name"] = pending_name

                append_operation_history(
                    stage="步骤二入库匹配",
                    action="执行匹配",
                    detail={
                        "source_file": step2_source_label,
                        "inbound_file": inbound_source_label,
                        "input_rows": len(df_step2_source),
                        "output_rows": len(df_step2_matched),
                        "inbound_rows": len(df_step2_inbound),
                        "pending_rows": len(df_step2_pending),
                        "alignment_can_compare": report_step2.get("can_compare"),
                        "alignment_ok": report_step2.get("ok") if report_step2.get("can_compare") else None,
                        "alignment_missing_rows": report_step2.get("missing_rows"),
                        "alignment_extra_rows": report_step2.get("extra_rows"),
                        "artifacts": [p for p in [inbound_artifact, pending_artifact] if p],
                    }
                )

                st.success("✅ 步骤二匹配完成。可直接在下方步骤三使用“已入库表”启动 AI 审核。")

            except ValueError as ve:
                st.error(f"❌ 步骤二匹配失败，缺少必要列：\n\n{ve}")
            except Exception as e:
                st.error(f"❌ 步骤二匹配失败：{e}")
                with st.expander("查看错误详情（开发用）"):
                    st.code(traceback.format_exc())

    cached_step2_matched = st.session_state.get("step2_matched_df")
    cached_step2_inbound = st.session_state.get("step2_inbound_df")
    cached_step2_pending = st.session_state.get("step2_not_inbound_df")
    if isinstance(cached_step2_matched, pd.DataFrame) and isinstance(cached_step2_inbound, pd.DataFrame) and isinstance(cached_step2_pending, pd.DataFrame):
        st.caption(f"当前会话缓存的步骤二已入库数据：{len(cached_step2_inbound)} 条。")

        c1, c2, c3 = st.columns(3)
        c1.metric("步骤二匹配总行数", len(cached_step2_matched))
        c2.metric("已入库（可进AI）", len(cached_step2_inbound))
        c3.metric("未入库（待跟进）", len(cached_step2_pending))

        render_preview_dataframe(
            cached_step2_inbound,
            title="✅ 已入库（可进入AI复核）",
            key_prefix="tab2_cached_inbound",
            default_rows=100,
            height=340,
            expanded=False,
        )
        render_preview_dataframe(
            cached_step2_pending,
            title="⚠️ 未入库（需继续跟进）",
            key_prefix="tab2_cached_pending",
            default_rows=100,
            height=340,
            expanded=False,
        )

        b_inbound = st.session_state.get("step2_inbound_bytes")
        b_pending = st.session_state.get("step2_pending_bytes")
        inbound_name = st.session_state.get("step2_inbound_name", f"{now_ts()}_入库匹配通过_待AI复核.xlsx")
        pending_name = st.session_state.get("step2_pending_name", f"{now_ts()}_未入库待跟进.xlsx")

        if not b_inbound or not b_pending:
            shot_col_step2 = find_first_existing_column(cached_step2_matched, COL_SCREENSHOT_CANDIDATES)
            hyperlink_cols_inbound = [shot_col_step2] if shot_col_step2 and shot_col_step2 in cached_step2_inbound.columns else None
            hyperlink_cols_pending = [shot_col_step2] if shot_col_step2 and shot_col_step2 in cached_step2_pending.columns else None
            b_inbound = df_to_excel_bytes(cached_step2_inbound, sheet_name="已入库", hyperlink_cols=hyperlink_cols_inbound)
            b_pending = df_to_excel_bytes(cached_step2_pending, sheet_name="未入库", hyperlink_cols=hyperlink_cols_pending)

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "⬇️ 下载：已入库表（待AI复核）",
                data=b_inbound,
                file_name=inbound_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with dl2:
            st.download_button(
                "⬇️ 下载：未入库表（待跟进）",
                data=b_pending,
                file_name=pending_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        report_step2_cached = st.session_state.get("step2_alignment_report")
        if isinstance(report_step2_cached, dict):
            render_alignment_report(report_step2_cached, title="步骤二源数据 vs 匹配结果一致性校验")

    st.divider()

    # -------------------------------
    # 步骤三（B）：AI 复核（多图）
    # -------------------------------
    st.subheader("步骤三（B）：AI 复核（多图）")
    st.info("推荐：直接使用步骤二产出的【已入库表】进行 AI 复核。")

    left, right = st.columns([2, 1])
    with left:
        uploaded_2 = st.file_uploader(
            "上传“确认入库后的正常表”（.xlsx / .xls / .csv）",
            type=["xlsx", "xls", "csv"],
            key="tab3_uploader"
        )
    with right:
        use_step2_df = st.checkbox(
            "直接使用步骤二的【已入库表】（推荐）",
            value=True,
            key="use_step2_df_for_ai"
        )

    if "active_ai_task_id" not in st.session_state:
        st.session_state["active_ai_task_id"] = ""
    if "ai_task_auto_refresh" not in st.session_state:
        st.session_state["ai_task_auto_refresh"] = True

    df_in = None
    file_bytes_2 = b""
    source_file_label = ""
    create_ready = False
    col_amount2 = ""
    col_shot2 = ""

    if use_step2_df and isinstance(st.session_state.get("step2_inbound_df"), pd.DataFrame):
        df_in = st.session_state["step2_inbound_df"].copy()
        source_file_label = "步骤二已入库表（会话）"
        st.success("已载入步骤二已入库表（当前会话）。")
        render_preview_dataframe(
            df_in,
            title="待 AI 审核数据预览（来自步骤二）",
            key_prefix="tab3_step2_session_df",
            default_rows=50,
            height=360,
            expanded=False,
        )
    elif uploaded_2 is not None:
        try:
            file_bytes_2 = get_uploaded_bytes(uploaded_2)
            df_in = read_table(uploaded_2)
            source_file_label = uploaded_2.name
            render_preview_dataframe(
                df_in,
                title="已读取待 AI 审核数据预览",
                key_prefix="tab3_uploaded_df",
                default_rows=50,
                height=360,
                expanded=False,
            )
        except Exception as e:
            st.error(f"❌ 文件读取失败：{e}")
            df_in = None

    current_task = None
    active_task_id = str(st.session_state.get("active_ai_task_id", "")).strip()
    if active_task_id:
        current_task = load_ai_task_state(active_task_id)
    if current_task is None:
        recovered_task = load_latest_ai_task_state(prefer_active=True)
        if recovered_task is not None:
            st.session_state["active_ai_task_id"] = recovered_task.get("task_id", "")
            current_task = recovered_task

    if current_task is not None:
        current_task = finalize_ai_task_if_needed(current_task)

    try:
        if df_in is not None and (not df_in.empty):
            required2 = {"退回运费金额": COL_AMOUNT_CANDIDATES, "寄回运费截图": COL_SCREENSHOT_CANDIDATES}
            matched2 = ensure_required_columns(df_in, required2)
            col_amount2 = matched2["退回运费金额"]
            col_shot2 = matched2["寄回运费截图"]

            # 若上传的是 Excel，抽取 screenshot 列 URL（兼容超链接/公式/tooltip/批注）
            if uploaded_2 is not None and uploaded_2.name.lower().endswith((".xlsx", ".xls")):
                df_in = attach_hyperlink_helper_column(df_in, file_bytes_2, col_shot2)
            create_ready = True

        if (df_in is None or df_in.empty) and current_task is None:
            st.info("请先在步骤二完成入库匹配并使用已入库表，或手动上传确认入库后的正常表。")

        ctrl1, ctrl2, ctrl3, ctrl4 = st.columns(4)
        with ctrl1:
            start_task = st.button(
                "🚀 创建并启动 AI 后台任务",
                type="primary",
                key="start_ai_task_background",
                disabled=not create_ready
            )
        with ctrl2:
            pause_task = st.button(
                "⏸️ 暂停任务",
                key="pause_ai_task",
                disabled=not (isinstance(current_task, dict) and current_task.get("status") == AI_TASK_STATUS_RUNNING)
            )
        with ctrl3:
            resume_task = st.button(
                "▶️ 继续任务",
                key="resume_ai_task",
                disabled=not (isinstance(current_task, dict) and current_task.get("status") in (AI_TASK_STATUS_PAUSED, AI_TASK_STATUS_ERROR))
            )
        with ctrl4:
            refresh_task = st.button("🔄 刷新任务状态", key="refresh_ai_task_status")

        st.checkbox("任务运行时自动刷新进度（每秒）", key="ai_task_auto_refresh")

        if start_task:
            if not api_key_input:
                st.warning("未检测到 DashScope API Key：请在侧边栏输入或配置环境变量 DASHSCOPE_API_KEY。")
            elif isinstance(current_task, dict) and current_task.get("status") == AI_TASK_STATUS_RUNNING:
                st.warning("当前已有运行中的 AI 任务，请先暂停后再创建新任务。")
            elif not create_ready:
                st.warning("当前数据未就绪，无法创建 AI 任务。")
            else:
                total_rows = min(len(df_in), int(max_ai_rows))
                new_task = create_ai_task_state(
                    df_source=df_in,
                    source_file=source_file_label,
                    col_amount=col_amount2,
                    col_shot=col_shot2,
                    total_rows=total_rows,
                    model_name=model_name,
                    max_images=int(max_images_per_row),
                    min_interval_sec=float(min_interval_sec),
                    max_retries=int(max_retries),
                    backoff_base_sec=float(backoff_base_sec),
                )
                st.session_state["active_ai_task_id"] = new_task.get("task_id", "")
                append_operation_history(
                    stage="步骤三AI复核",
                    action="创建AI任务",
                    detail={
                        "task_id": new_task.get("task_id", ""),
                        "source_file": source_file_label,
                        "input_rows": len(df_in),
                        "processed_rows": 0,
                        "max_rows": total_rows,
                        "model": model_name,
                    }
                )
                if start_ai_task_worker(new_task.get("task_id", ""), api_key_input):
                    st.success(f"✅ AI 后台任务已启动：{new_task.get('task_id', '')}")
                else:
                    st.error("❌ AI 任务启动失败，请检查任务状态后重试。")
                current_task = load_ai_task_state(new_task.get("task_id", ""))

        if pause_task and isinstance(current_task, dict):
            current_task["status"] = AI_TASK_STATUS_PAUSED
            current_task["worker_token"] = ""
            save_ai_task_state(current_task)
            append_operation_history(
                stage="步骤三AI复核",
                action="暂停AI任务",
                detail={
                    "task_id": current_task.get("task_id", ""),
                    "processed_rows": int(current_task.get("next_idx", 0)),
                    "max_rows": int(current_task.get("total", 0)),
                }
            )
            st.info("任务已暂停。")
            current_task = load_ai_task_state(current_task.get("task_id", ""))

        if resume_task and isinstance(current_task, dict):
            if not api_key_input:
                st.warning("继续任务需要 DashScope API Key，请先在侧边栏输入。")
            else:
                if start_ai_task_worker(current_task.get("task_id", ""), api_key_input):
                    append_operation_history(
                        stage="步骤三AI复核",
                        action="继续AI任务",
                        detail={
                            "task_id": current_task.get("task_id", ""),
                            "processed_rows": int(current_task.get("next_idx", 0)),
                            "max_rows": int(current_task.get("total", 0)),
                        }
                    )
                    st.success("任务已继续执行。")
                else:
                    st.error("❌ 任务继续失败，请重试。")
                current_task = load_ai_task_state(current_task.get("task_id", ""))

        if refresh_task and isinstance(current_task, dict):
            current_task = load_ai_task_state(current_task.get("task_id", ""))

        if isinstance(current_task, dict):
            current_task = finalize_ai_task_if_needed(current_task)
            summary = summarize_ai_task(current_task)
            frames = split_ai_task_frames(current_task)

            st.markdown(f"#### 🧠 当前任务：`{current_task.get('task_id', '')}`（{task_status_label(current_task.get('status', ''))}）")
            if current_task.get("error_message"):
                st.warning(f"任务提示：{current_task.get('error_message')}")

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("计划处理", summary["total"])
            m2.metric("已处理", summary["processed_rows"])
            m3.metric("未处理", summary["pending_rows"])
            m4.metric("处理正常", summary["ok_rows"])
            m5.metric("处理异常", summary["bad_rows"])

            total = max(1, int(summary["total"]))
            progress_ratio = min(1.0, float(summary["processed_rows"]) / float(total))
            st.progress(progress_ratio)
            st.caption(
                f"进度：{summary['processed_rows']}/{summary['total']} | "
                f"创建时间：{current_task.get('created_at', '')} | "
                f"最近更新：{current_task.get('updated_at', '')}"
            )

            col_shot_task = str(current_task.get("col_shot", ""))
            hyperlink_ok = [col_shot_task] if col_shot_task and col_shot_task in frames["ok"].columns else None
            hyperlink_bad = [col_shot_task] if col_shot_task and col_shot_task in frames["bad"].columns else None
            hyperlink_pending = [col_shot_task] if col_shot_task and col_shot_task in frames["pending"].columns else None

            b_ok_partial = df_to_excel_bytes(frames["ok"], sheet_name="AI已处理正常", hyperlink_cols=hyperlink_ok)
            b_bad_partial = df_to_excel_bytes(frames["bad"], sheet_name="AI已处理异常", hyperlink_cols=hyperlink_bad)
            b_pending_partial = df_to_excel_bytes(frames["pending"], sheet_name="AI未处理", hyperlink_cols=hyperlink_pending)

            dl1, dl2, dl3 = st.columns(3)
            with dl1:
                st.download_button(
                    "⬇️ 下载已处理正常",
                    data=b_ok_partial,
                    file_name=f"{now_ts()}_{current_task.get('task_id', '')}_AI已处理正常.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with dl2:
                st.download_button(
                    "⬇️ 下载已处理异常",
                    data=b_bad_partial,
                    file_name=f"{now_ts()}_{current_task.get('task_id', '')}_AI已处理异常.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with dl3:
                st.download_button(
                    "⬇️ 下载未处理部分",
                    data=b_pending_partial,
                    file_name=f"{now_ts()}_{current_task.get('task_id', '')}_AI未处理.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with st.expander("查看已处理明细（最近100条）", expanded=False):
                if frames["processed"].empty:
                    st.info("暂无已处理记录。")
                else:
                    st.dataframe(frames["processed"].tail(100), use_container_width=True, height=360)

            with st.expander("查看未处理明细（最多100条）", expanded=False):
                if frames["pending"].empty:
                    st.info("当前没有未处理记录。")
                else:
                    st.dataframe(frames["pending"].head(100), use_container_width=True, height=360)

            align_rep = current_task.get("alignment_report")
            if isinstance(align_rep, dict):
                render_alignment_report(align_rep, title="步骤三源数据 vs AI结果一致性校验")

            if current_task.get("status") == AI_TASK_STATUS_RUNNING and st.session_state.get("ai_task_auto_refresh", True):
                time.sleep(1.0)
                st.rerun()

    except ValueError as ve:
        st.error(f"❌ 表格缺少必要列，请检查：\n\n{ve}")
    except Exception as e:
        st.error(f"❌ AI 审核流程失败：{e}")
        with st.expander("查看错误详情（开发用）"):
            st.code(traceback.format_exc())


# =============================================================================
# Tab1：清洗与规则初筛
# =============================================================================
with main_tabs[1]:
    st.subheader("步骤一：上传班牛《退运费自助登记表》并进行规则初筛")

    uploaded_1 = st.file_uploader("上传《退运费自助登记表》（.xlsx / .xls / .csv）", type=["xlsx", "xls", "csv"], key="tab1_uploader")

    if uploaded_1 is None:
        st.info("请先上传班牛导出的登记表。")
    else:
        try:
            file_bytes_1 = get_uploaded_bytes(uploaded_1)
            df_raw = read_table(uploaded_1)

            if df_raw.empty:
                st.warning("读取到的表格为空，请检查文件内容。")
            else:
                render_preview_dataframe(
                    df_raw,
                    title="已读取数据预览",
                    key_prefix="tab1_uploaded_raw",
                    default_rows=50,
                    height=420,
                    expanded=False,
                )

                required = {
                    "退回运费金额": COL_AMOUNT_CANDIDATES,
                    "支付宝账号": COL_ALIPAY_ACCOUNT_CANDIDATES,
                    "支付宝实名": COL_ALIPAY_NAME_CANDIDATES,
                    "退回物流单号": COL_LOGISTICS_NO_CANDIDATES,
                }
                matched_cols = ensure_required_columns(df_raw, required)

                col_amount = matched_cols["退回运费金额"]
                col_account = matched_cols["支付宝账号"]
                col_name = matched_cols["支付宝实名"]
                col_lno = matched_cols["退回物流单号"]

                shot_col = find_first_existing_column(df_raw, COL_SCREENSHOT_CANDIDATES)

                # ✅ Excel 截图列 URL 抽取（兼容 hyperlink / 公式 / tooltip / 批注）
                if shot_col and uploaded_1.name.lower().endswith((".xlsx", ".xls")):
                    df_raw = attach_hyperlink_helper_column(df_raw, file_bytes_1, shot_col)

                # 校验
                df = df_raw.copy()
                validation_results = [
                    validate_row(
                        amount=amount,
                        alipay_account=account,
                        alipay_name=name,
                        logistics_no=logistics_no,
                    )
                    for amount, account, name, logistics_no in zip(
                        df[col_amount].tolist(),
                        df[col_account].tolist(),
                        df[col_name].tolist(),
                        df[col_lno].tolist(),
                    )
                ]
                flags = [ok for ok, _ in validation_results]
                reasons = [reason for _, reason in validation_results]

                df[COL_ABNORMAL_REASON] = reasons
                valid_mask = pd.Series(flags, index=df.index)
                df_normal = df[valid_mask].drop(columns=[COL_ABNORMAL_REASON], errors="ignore").copy()
                df_abnormal = df[~valid_mask].copy()
                report_step1 = compare_source_and_processed(df_raw, df, stage_name="步骤一清洗")

                c1, c2, c3 = st.columns(3)
                c1.metric("总行数", len(df))
                c2.metric("正常（可继续反查）", len(df_normal))
                c3.metric("异常（需回访）", len(df_abnormal))

                render_preview_dataframe(
                    df_normal,
                    title="✅ 正常表",
                    key_prefix="tab1_normal_df",
                    default_rows=100,
                    height=420,
                    expanded=False,
                )
                render_preview_dataframe(
                    df_abnormal,
                    title="⚠️ 异常表",
                    key_prefix="tab1_abnormal_df",
                    default_rows=100,
                    height=420,
                    expanded=False,
                )

                # session 供 Tab2 直接用
                st.session_state["tab1_normal_df"] = df_normal
                st.session_state["tab1_abnormal_df"] = df_abnormal

                # 导出：截图列保持“预览/浏览”等原文字，但可点击
                hyperlink_cols_normal = [shot_col] if shot_col and shot_col in df_normal.columns else None
                hyperlink_cols_abnormal = [shot_col] if shot_col and shot_col in df_abnormal.columns else None

                ts = now_ts()
                normal_name = f"{ts}_清洗正常可继续反查.xlsx"
                abnormal_name = f"{ts}_退运费信息异常需回访.xlsx"

                b1 = df_to_excel_bytes(df_normal, sheet_name="正常", hyperlink_cols=hyperlink_cols_normal)
                b2 = df_to_excel_bytes(df_abnormal, sheet_name="异常", hyperlink_cols=hyperlink_cols_abnormal)

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        "⬇️ 下载正常表（可继续反查）",
                        data=b1,
                        file_name=normal_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col_dl2:
                    st.download_button(
                        "⬇️ 下载异常表（需回访）",
                        data=b2,
                        file_name=abnormal_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                render_alignment_report(report_step1, title="步骤一源数据 vs 清洗结果一致性校验")

                step1_history_key = f"{uploaded_1.name}|{len(file_bytes_1)}|{len(df)}|{len(df.columns)}"
                if st.session_state.get("step1_last_history_key") != step1_history_key:
                    normal_artifact = save_artifact_bytes("step1_normal", normal_name, b1)
                    abnormal_artifact = save_artifact_bytes("step1_abnormal", abnormal_name, b2)
                    append_operation_history(
                        stage="步骤一清洗",
                        action="执行清洗",
                        detail={
                            "source_file": uploaded_1.name,
                            "input_rows": len(df_raw),
                            "output_rows": len(df),
                            "normal_rows": len(df_normal),
                            "abnormal_rows": len(df_abnormal),
                            "alignment_can_compare": report_step1.get("can_compare"),
                            "alignment_ok": report_step1.get("ok") if report_step1.get("can_compare") else None,
                            "alignment_missing_rows": report_step1.get("missing_rows"),
                            "alignment_extra_rows": report_step1.get("extra_rows"),
                            "artifacts": [p for p in [normal_artifact, abnormal_artifact] if p],
                        }
                    )
                    st.session_state["step1_last_history_key"] = step1_history_key

        except ValueError as ve:
            st.error(f"❌ 表格缺少必要列，请检查后重新导出/上传：\n\n{ve}")
        except Exception as e:
            st.error(f"❌ 处理失败：{e}")
            with st.expander("查看错误详情（开发用）"):
                st.code(traceback.format_exc())
