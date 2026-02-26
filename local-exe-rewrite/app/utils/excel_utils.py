from __future__ import annotations

import math
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, unquote, urlparse

import pandas as pd
from openpyxl import load_workbook

from app.core.constants import (
    HYPERLINK_SUFFIX,
    IDENTIFIER_COLUMN_KEYWORDS,
    IMAGE_EXTENSIONS,
    REGEX_EXCEL_HYPERLINK_FORMULA,
    REGEX_EXCEL_URL_FALLBACK,
    REGEX_PREVIEW_SPLIT,
    REGEX_SCI_NUMBER,
    REGEX_URL_GENERIC,
    REGEX_URL_IN_PARENS,
)


def safe_strip_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def read_table(file_bytes: bytes, filename: str) -> pd.DataFrame:
    if not file_bytes:
        return pd.DataFrame()

    name = str(filename or "").lower()
    bio = BytesIO(file_bytes)

    try:
        if name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(
                bio,
                engine="openpyxl",
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
        elif name.endswith(".csv"):
            try:
                df = pd.read_csv(bio, encoding="utf-8", dtype=str, keep_default_na=False, na_filter=False)
            except UnicodeDecodeError:
                bio.seek(0)
                df = pd.read_csv(bio, encoding="gbk", dtype=str, keep_default_na=False, na_filter=False)
        else:
            raise ValueError("仅支持 .xlsx / .xls / .csv")
    except Exception as exc:
        raise RuntimeError(f"文件读取失败: {exc}") from exc

    return safe_strip_columns(df)


def _dedupe_preserve_order(values: List[str], max_items: Optional[int] = None) -> List[str]:
    seen = set()
    out: List[str] = []
    for value in values:
        item = str(value or "").strip()
        if not item or item in seen:
            continue
        out.append(item)
        seen.add(item)
        if max_items is not None and len(out) >= max_items:
            break
    return out


def extract_urls_from_cell(cell_value: Any) -> List[str]:
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    raw = str(cell_value).strip()
    if not raw:
        return []

    urls: List[str] = []
    urls.extend(REGEX_URL_IN_PARENS.findall(raw))
    urls.extend(REGEX_URL_GENERIC.findall(raw))
    return _dedupe_preserve_order(urls)


@lru_cache(maxsize=4096)
def _normalize_preview_url_cached(url: str) -> Tuple[str, ...]:
    if not url:
        return tuple()
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        if "url" in qs:
            raw = unquote(qs["url"][0])
            parts = REGEX_PREVIEW_SPLIT.split(raw)
            extracted = [p.strip() for p in parts if p.strip().startswith("http")]
            if extracted:
                return tuple(extracted)
    except Exception:
        pass
    return (url,)


def normalize_preview_url(url: str) -> List[str]:
    return list(_normalize_preview_url_cached(str(url or "").strip()))


def pick_image_urls(urls: List[str], max_images: int = 4) -> List[str]:
    if not urls:
        return []

    expanded: List[str] = []
    for url in urls:
        expanded.extend(normalize_preview_url(url))

    image_urls = [url for url in expanded if str(url).lower().endswith(IMAGE_EXTENSIONS)]
    out = _dedupe_preserve_order(image_urls, max_items=max_images)

    if not out:
        http_urls = [url for url in expanded if str(url).startswith("http")]
        out = _dedupe_preserve_order(http_urls, max_items=max_images)

    return out


@lru_cache(maxsize=8192)
def _extract_image_urls_from_text(raw_text: str, max_images: int) -> Tuple[str, ...]:
    urls = extract_urls_from_cell(raw_text)
    img_urls = pick_image_urls(urls, max_images=max_images)
    if img_urls:
        return tuple(img_urls)

    if raw_text.startswith("http"):
        expanded = normalize_preview_url(raw_text)
        image_candidates = [url for url in expanded if url.lower().endswith(IMAGE_EXTENSIONS)]
        if image_candidates:
            return tuple(image_candidates[:max_images])
        return tuple(expanded[:max_images])
    return tuple()


def extract_image_urls_from_cell_value(cell_value: Any, max_images: int = 4) -> List[str]:
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    raw_text = str(cell_value).strip()
    if not raw_text:
        return []
    return list(_extract_image_urls_from_text(raw_text, int(max_images)))


def extract_hyperlinks_from_excel(
    file_bytes: bytes,
    target_header: str,
    n_rows: Optional[int] = None,
) -> List[Optional[str]]:
    if not file_bytes:
        return []

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        ws = wb.worksheets[0]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            headers.append(str(value).strip() if value is not None else "")

        if target_header not in headers:
            return []

        col_idx = headers.index(target_header) + 1
        end_row = ws.max_row if n_rows is None else 1 + int(n_rows)

        links: List[Optional[str]] = []
        for row in range(2, end_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            url = None

            if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                url = str(cell.hyperlink.target).strip()

            if not url:
                value = cell.value
                if isinstance(value, str):
                    m = REGEX_EXCEL_HYPERLINK_FORMULA.search(value)
                    if m:
                        url = m.group(1).strip()

            if not url:
                tip = None
                try:
                    tip = getattr(cell.hyperlink, "tooltip", None) if cell.hyperlink else None
                except Exception:
                    tip = None
                if isinstance(tip, str):
                    m = REGEX_EXCEL_URL_FALLBACK.search(tip)
                    if m:
                        url = m.group(1).strip()

            if not url and cell.comment and isinstance(cell.comment.text, str):
                m = REGEX_EXCEL_URL_FALLBACK.search(cell.comment.text)
                if m:
                    url = m.group(1).strip()

            links.append(url if url else None)

        return links
    except Exception:
        return []


def attach_hyperlink_helper_column(df: pd.DataFrame, file_bytes: bytes, screenshot_col: str) -> pd.DataFrame:
    out = df.copy()
    if not file_bytes or out.empty:
        return out

    links = extract_hyperlinks_from_excel(file_bytes, screenshot_col, n_rows=len(out))
    if len(links) < len(out):
        links += [None] * (len(out) - len(links))
    elif len(links) > len(out):
        links = links[: len(out)]

    out[screenshot_col + HYPERLINK_SUFFIX] = links
    return out


def _is_identifier_column(col_name: str) -> bool:
    name = str(col_name)
    return any(key in name for key in IDENTIFIER_COLUMN_KEYWORDS)


def _normalize_scientific_text(text: str) -> str:
    value = str(text).strip()
    if not value or not REGEX_SCI_NUMBER.match(value):
        return value
    try:
        d = Decimal(value)
    except InvalidOperation:
        return value

    if d == d.to_integral_value():
        return format(d.quantize(Decimal("1")), "f")
    out = format(d, "f").rstrip("0").rstrip(".")
    return out if out else "0"


def _normalize_identifier_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return format(value, ".0f")
        out = format(value, "f").rstrip("0").rstrip(".")
        return out if out else "0"
    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    if not text:
        return ""
    return _normalize_scientific_text(text)


def df_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "sheet1",
    hyperlink_cols: Optional[List[str]] = None,
) -> bytes:
    export_df = df.copy()
    identifier_cols = [col for col in export_df.columns if _is_identifier_column(col)]
    for col in identifier_cols:
        export_df[col] = export_df[col].map(_normalize_identifier_cell)

    link_targets: Dict[str, List[Optional[str]]] = {}

    if hyperlink_cols:
        for col in hyperlink_cols:
            helper_col = col + HYPERLINK_SUFFIX
            if helper_col in export_df.columns:
                link_targets[col] = export_df[helper_col].tolist()
                export_df.drop(columns=[helper_col], inplace=True, errors="ignore")

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)

    if not hyperlink_cols:
        return bio.getvalue()

    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            header_map[str(value).strip()] = col

    for col in identifier_cols:
        if col not in header_map:
            continue
        col_idx = header_map[col]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            cell.number_format = "@"

    for col in hyperlink_cols:
        if col not in header_map:
            continue
        col_idx = header_map[col]
        targets = link_targets.get(col, [])

        for row in range(2, ws.max_row + 1):
            df_idx = row - 2
            cell = ws.cell(row=row, column=col_idx)

            target = None
            if targets and df_idx < len(targets):
                target = targets[df_idx]

            if not target:
                text = "" if cell.value is None else str(cell.value).strip()
                urls = extract_urls_from_cell(text)
                imgs = pick_image_urls(urls, max_images=1)
                target = imgs[0] if imgs else None
                if not target and text.startswith("http"):
                    expanded = normalize_preview_url(text)
                    target = expanded[0] if expanded else text

            if target and str(target).startswith("http"):
                cell.hyperlink = str(target).strip()
                cell.style = "Hyperlink"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def df_to_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    if df.empty:
        return []
    safe_df = df.astype(object).where(pd.notna(df), None)
    return safe_df.to_dict(orient="records")
