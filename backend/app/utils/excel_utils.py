# app/utils/excel_utils.py
import math
import pandas as pd
from io import BytesIO
from typing import List, Tuple, Optional, Any, Dict
from functools import lru_cache
from urllib.parse import unquote, urlparse, parse_qs
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook

from app.core.constants import (
    REGEX_URL_IN_PARENS, REGEX_URL_GENERIC, REGEX_PREVIEW_SPLIT,
    IMAGE_EXTENSIONS, REGEX_EXCEL_HYPERLINK_FORMULA, REGEX_EXCEL_URL_FALLBACK,
    HYPERLINK_SUFFIX, IDENTIFIER_COLUMN_KEYWORDS, REGEX_SCI_NUMBER
)

def safe_strip_columns(df: pd.DataFrame) -> pd.DataFrame:
    """【强防坑要求】读取后立刻 strip 列名"""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def read_table(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """读取 xlsx/xls/csv，并在第一时间 strip 列名"""
    if not file_bytes:
        return pd.DataFrame()

    filename = filename.lower()
    bio = BytesIO(file_bytes)
    
    try:
        if filename.endswith((".xlsx", ".xls")):
            # 关键字段（订单号/账号/单号）需保持文本，避免科学计数法和精度风险
            df = pd.read_excel(
                bio,
                engine="openpyxl",
                dtype=str,
                keep_default_na=False,
                na_filter=False
            )
        elif filename.endswith(".csv"):
            try:
                df = pd.read_csv(
                    bio,
                    encoding="utf-8",
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False
                )
            except UnicodeDecodeError:
                bio.seek(0)
                df = pd.read_csv(
                    bio,
                    encoding="gbk",
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False
                )
        else:
            raise ValueError("仅支持 .xlsx / .xls / .csv")
    except Exception as e:
        raise RuntimeError(f"文件读取失败：{e}")

    return safe_strip_columns(df)

def _dedupe_preserve_order(values: List[str], max_items: Optional[int] = None) -> List[str]:
    seen = set()
    out: List[str] = []
    for v in values:
        item = v.strip()
        if not item or item in seen:
            continue
        out.append(item)
        seen.add(item)
        if max_items is not None and len(out) >= max_items:
            break
    return out

def extract_urls_from_cell(cell_value: Any) -> List[str]:
    """从单元格文本中提取 URL（支持 markdown/裸链）"""
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    urls: List[str] = []
    urls.extend(REGEX_URL_IN_PARENS.findall(s))
    urls.extend(REGEX_URL_GENERIC.findall(s))
    return _dedupe_preserve_order(urls)

@lru_cache(maxsize=4096)
def _normalize_preview_url_cached(url: str) -> Tuple[str, ...]:
    if not url:
        return tuple()
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        if "url" in qs:
            raw = unquote(qs["url"][0])
            parts = REGEX_PREVIEW_SPLIT.split(raw)
            extracted = [p.strip() for p in parts if p.strip().startswith("http")]
            if extracted:
                return tuple(extracted)
    except Exception:
        pass
    return (url,)

def normalize_preview_url(url: str) -> List[str]:
    if not url:
        return []
    return list(_normalize_preview_url_cached(str(url).strip()))

def pick_image_urls(urls: List[str], max_images: int = 4) -> List[str]:
    """多图：从单元格解析出的链接拆出多图，优先提取直链图片"""
    if not urls:
        return []
    expanded: List[str] = []
    for u in urls:
        expanded.extend(normalize_preview_url(u))

    imgs = [u for u in expanded if u.lower().endswith(IMAGE_EXTENSIONS)]
    out = _dedupe_preserve_order(imgs, max_items=max_images)

    if not out:
        http_urls = [u for u in expanded if u.startswith("http")]
        out = _dedupe_preserve_order(http_urls, max_items=max_images)

    return out

@lru_cache(maxsize=8192)
def _extract_image_urls_from_text(raw_text: str, max_images: int) -> Tuple[str, ...]:
    urls = extract_urls_from_cell(raw_text)
    img_urls = pick_image_urls(urls, max_images=max_images)
    if img_urls:
        return tuple(img_urls)

    if raw_text.startswith("http"):
        expanded = normalize_preview_url(raw_text)
        image_candidates = [u for u in expanded if u.lower().endswith(IMAGE_EXTENSIONS)]
        if image_candidates:
            return tuple(image_candidates[:max_images])
        return tuple(expanded[:max_images])
    return tuple()

def extract_image_urls_from_cell_value(cell_value: Any, max_images: int = 4) -> List[str]:
    if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
        return []
    raw_text = str(cell_value).strip()
    if not raw_text:
        return []
    return list(_extract_image_urls_from_text(raw_text, int(max_images)))

def extract_hyperlinks_from_excel(file_bytes: bytes, target_header: str, n_rows: Optional[int] = None) -> List[Optional[str]]:
    """
    【命脉代码：严禁修改结构】
    从 Excel 中提取指定列每行的链接 URL，兼容：原生超链接、公式、tooltip、批注。
    """
    if not file_bytes:
        return []
    try:
        # 关键：data_only=False 才能拿到公式本体
        wb = load_workbook(BytesIO(file_bytes), data_only=False)
        ws = wb.worksheets[0]

        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            headers.append(str(v).strip() if v is not None else "")

        if target_header not in headers:
            return []

        col_idx = headers.index(target_header) + 1
        end_row = ws.max_row if n_rows is None else 1 + int(n_rows)

        links: List[Optional[str]] = []
        for r in range(2, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            url = None

            # 1) 原生 hyperlink
            if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                url = str(cell.hyperlink.target).strip()

            # 2) HYPERLINK公式
            if not url:
                v = cell.value
                if isinstance(v, str):
                    m = REGEX_EXCEL_HYPERLINK_FORMULA.search(v)
                    if m:
                        url = m.group(1).strip()

            # 3) tooltip / comment 兜底
            if not url:
                tip = None
                try:
                    tip = getattr(cell.hyperlink, "tooltip", None) if cell.hyperlink else None
                except Exception:
                    pass
                if isinstance(tip, str):
                    m = REGEX_EXCEL_URL_FALLBACK.search(tip)
                    if m:
                        url = m.group(1).strip()

            if not url and cell.comment and isinstance(cell.comment.text, str):
                m = REGEX_EXCEL_URL_FALLBACK.search(cell.comment.text)
                if m:
                    url = m.group(1).strip()

            links.append(url if url else None)

        return links
    except Exception:
        return []

def attach_hyperlink_helper_column(df: pd.DataFrame, file_bytes: bytes, screenshot_col: str) -> pd.DataFrame:
    """挂载超链接辅助列"""
    df = df.copy()
    if not file_bytes or df.empty:
        return df

    links = extract_hyperlinks_from_excel(file_bytes, screenshot_col, n_rows=len(df))

    if len(links) < len(df):
        links = links + [None] * (len(df) - len(links))
    elif len(links) > len(df):
        links = links[:len(df)]

    df[screenshot_col + HYPERLINK_SUFFIX] = links
    return df

def _is_identifier_column(col_name: str) -> bool:
    name = str(col_name)
    return any(k in name for k in IDENTIFIER_COLUMN_KEYWORDS)

def _normalize_scientific_text(s: str) -> str:
    v = s.strip()
    if not v or not REGEX_SCI_NUMBER.match(v):
        return v
    try:
        d = Decimal(v)
    except InvalidOperation:
        return v
    if d == d.to_integral_value():
        return format(d.quantize(Decimal("1")), "f")
    out = format(d, "f").rstrip("0").rstrip(".")
    return out if out else "0"

def _normalize_identifier_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return format(value, ".0f")
        out = format(value, "f").rstrip("0").rstrip(".")
        return out if out else "0"
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s:
        return ""
    return _normalize_scientific_text(s)

def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "sheet1", hyperlink_cols: Optional[List[str]] = None) -> bytes:
    """
    【命脉代码：严禁修改结构】
    DataFrame -> Excel bytes，写回超链接，保留原文字（如“预览/浏览”）但让整格可点击。
    """
    df_export = df.copy()
    identifier_cols = [c for c in df_export.columns if _is_identifier_column(c)]
    for col in identifier_cols:
        df_export[col] = df_export[col].map(_normalize_identifier_cell)

    link_targets: Dict[str, List[Optional[str]]] = {}

    if hyperlink_cols:
        for col in hyperlink_cols:
            helper_col = col + HYPERLINK_SUFFIX
            if helper_col in df_export.columns:
                link_targets[col] = df_export[helper_col].tolist()
                df_export.drop(columns=[helper_col], inplace=True, errors="ignore")

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)

    if not hyperlink_cols:
        return bio.getvalue()

    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            header_map[str(v).strip()] = c

    for col in identifier_cols:
        if col not in header_map:
            continue
        cidx = header_map[col]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=cidx)
            if cell.value is None:
                continue
            cell.value = str(cell.value)
            cell.number_format = "@"

    for col in hyperlink_cols:
        if col not in header_map:
            continue
        cidx = header_map[col]
        targets = link_targets.get(col, [])

        for r in range(2, ws.max_row + 1):
            df_idx = r - 2
            cell = ws.cell(row=r, column=cidx)

            target = None
            if targets and df_idx < len(targets):
                target = targets[df_idx]

            if not target:
                val = "" if cell.value is None else str(cell.value).strip()
                urls = extract_urls_from_cell(val)
                imgs = pick_image_urls(urls, max_images=1)
                target = imgs[0] if imgs else None
                if not target and val.startswith("http"):
                    expanded = normalize_preview_url(val)
                    target = expanded[0] if expanded else val

            if target and str(target).startswith("http"):
                cell.hyperlink = str(target).strip()
                cell.style = "Hyperlink"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()